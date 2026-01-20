from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib import messages
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User, Group
from django.core.mail import send_mail
from django.http import HttpResponse, JsonResponse
import random
from django.conf import settings
from datetime import datetime, timedelta
from django.utils import timezone
import time
from .models import (
    Student, Teacher, Course, Activity, Payment, Invoice, Event, 
    Timetable, Report, Examination, ExamSchedule, Settings, Profile, 
    Notice, Attendance, Message, Announcement, Class, Schedule, 
    MessageAttachment, AnnouncementAttachment, Department, Grade, 
    Behavior, Fee, AssignmentSubmission, CourseMaterial, Evaluation, Expense
)
from .forms import (
    StudentForm, TeacherForm, CourseForm, StudentProfileForm, 
    AssignmentSubmissionForm, TeacherProfileForm, StudentRegistrationForm, 
    PaymentForm, InvoiceForm, GradeForm, BehaviorForm, FeeForm, 
    ClassForm, SubjectForm, AttendanceForm, TeacherRegistrationForm, EventForm, ScheduleForm,
    ExaminationForm, ExamScheduleForm, StudentPasswordChangeForm
)
from django.db.models import Q, Avg
from django.db import models
import csv
import xlsxwriter
from reportlab.pdfgen import canvas
from django.contrib.auth.hashers import make_password, check_password
from django.core.exceptions import ValidationError
from django.db import models
from django.views.decorators.http import require_http_methods
from django.template.loader import render_to_string
import json
from django.db.models import Sum, Count, Avg
from django.views.decorators.cache import never_cache
from django.core.paginator import Paginator
from django.core.files.storage import default_storage
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
import tempfile
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        role = request.POST.get('role')
        
        try:
            if role == 'teacher':
                # Try to find teacher by teacher_id first
                try:
                    teacher = Teacher.objects.get(teacher_id=username)
                    user = teacher.user
                    # Authenticate using the user's username and password
                    authenticated_user = authenticate(request, username=user.username, password=password)
                    if authenticated_user is None:
                        messages.error(request, 'Invalid password.')
                        return render(request, 'accounts/login.html')
                    user = authenticated_user
                except Teacher.DoesNotExist:
                    # If not found by teacher_id, try to authenticate with username directly
                    user = authenticate(request, username=username, password=password)
                    if user is None:
                        messages.error(request, 'Invalid Teacher ID or username.')
                        return render(request, 'accounts/login.html')
                    # Verify that the authenticated user is a teacher
                    if not hasattr(user, 'teacher_profile'):
                        messages.error(request, 'This account is not registered as a teacher.')
                        return render(request, 'accounts/login.html')
            elif role == 'student':
                # Try to find student by student_id first
                try:
                    student = Student.objects.get(student_id=username)
                    user = student.user
                    # Authenticate using the user's username and password
                    authenticated_user = authenticate(request, username=user.username, password=password)
                    if authenticated_user is None:
                        messages.error(request, 'Invalid password.')
                        return render(request, 'accounts/login.html')
                    user = authenticated_user
                except Student.DoesNotExist:
                    # If not found by student_id, try to authenticate with username directly
                    user = authenticate(request, username=username, password=password)
                    if user is None:
                        messages.error(request, 'Invalid Student ID or username.')
                        return render(request, 'accounts/login.html')
                    # Verify that the authenticated user is a student
                    if not hasattr(user, 'student_profile'):
                        messages.error(request, 'This account is not registered as a student.')
                        return render(request, 'accounts/login.html')
            else:
                # For admin role, use the username directly
                user = authenticate(request, username=username, password=password)
            
            if user is not None:
                # Check if user has the appropriate role
                if role == 'admin' and user.is_staff:
                    login(request, user)
                    return redirect('dashboard')
                elif role == 'teacher' and hasattr(user, 'teacher_profile'):
                    if user.teacher_profile.is_active:
                        login(request, user)
                        return redirect('teacher_dashboard')
                    else:
                        messages.error(request, 'Your teacher account is not active. Please contact the administrator.')
                elif role == 'student' and hasattr(user, 'student_profile'):
                    if user.student_profile.status == 'active':
                        login(request, user)
                        return redirect('student_dashboard')
                    else:
                        messages.error(request, 'Your student account is not active. Please contact the administrator.')
                else:
                    messages.error(request, f'Invalid login credentials for {role}. Please make sure you selected the correct role.')
            else:
                messages.error(request, 'Invalid password. Please try again.')
        except Exception as e:
            messages.error(request, 'An error occurred during login. Please try again.')
            print(f"Login error: {str(e)}")  # For debugging purposes
    
    return render(request, 'accounts/login.html')

def register_view(request):
    if request.method == 'POST':
        role = request.POST.get('role', '')
        print(f"Registration attempt for role: {role}")  # Debug log
        
        try:
            if role == 'student':
                # Log student registration data
                print(f"Student registration data: {request.POST}")  # Debug log
                
                # Validate required fields
                required_fields = ['student_id', 'first_name', 'last_name', 'email', 'password1', 'password2']
                for field in required_fields:
                    if not request.POST.get(field):
                        print(f"Missing required field: {field}")  # Debug log
                        messages.error(request, f'{field.replace("_", " ").title()} is required.')
                        return render(request, 'accounts/register.html')
                
                # Check if student ID already exists
                student_id = request.POST.get('student_id')
                if Student.objects.filter(student_id=student_id).exists():
                    print(f"Student ID already exists: {student_id}")  # Debug log
                    messages.error(request, 'Student ID already exists.')
                    return render(request, 'accounts/register.html')
                
                # Check if username (student_id) already exists
                if User.objects.filter(username=student_id).exists():
                    print(f"Username already exists: {student_id}")  # Debug log
                    messages.error(request, 'This Student ID is already registered as a username.')
                    return render(request, 'accounts/register.html')
                
                # Check if email already exists
                email = request.POST.get('email')
                if User.objects.filter(email=email).exists():
                    print(f"Email already exists: {email}")  # Debug log
                    messages.error(request, 'Email already exists.')
                    return render(request, 'accounts/register.html')
                
                # Validate password
                if request.POST.get('password1') != request.POST.get('password2'):
                    print("Passwords do not match")  # Debug log
                    messages.error(request, 'Passwords do not match.')
                    return render(request, 'accounts/register.html')
                
                try:
                    # Create user account for student
                    user = User.objects.create_user(
                        username=student_id,
                        password=request.POST.get('password1'),
                        email=email,
                        first_name=request.POST.get('first_name'),
                        last_name=request.POST.get('last_name')
                    )
                    print(f"Created user account for student: {user.username}")  # Debug log
                except Exception as e:
                    print(f"Error creating user account: {str(e)}")  # Debug log
                    messages.error(request, f'Error creating user account: {str(e)}')
                    return render(request, 'accounts/register.html')
                
                try:
                    # Add user to student group
                    student_group = Group.objects.get_or_create(name='Student')[0]
                    user.groups.add(student_group)
                    print(f"Added user to student group")  # Debug log
                except Exception as e:
                    print(f"Error adding user to group: {str(e)}")  # Debug log
                    messages.error(request, f'Error adding user to group: {str(e)}')
                    return render(request, 'accounts/register.html')
                
                try:
                    # Create student profile with all required fields
                    Student.objects.create(
                        user=user,
                        student_id=student_id,
                        first_name=request.POST.get('first_name'),
                        last_name=request.POST.get('last_name'),
                        date_of_birth=request.POST.get('date_of_birth'),
                        gender=request.POST.get('gender'),
                        email=email,
                        phone_number=request.POST.get('phone_number', ''),
                        address=request.POST.get('address', ''),
                        grade_level=request.POST.get('grade_level', ''),
                        admission_date=datetime.now().date(),  # Using current date as admission date
                        guardian_name=request.POST.get('guardian_name', ''),
                        guardian_relation=request.POST.get('guardian_relation', ''),
                        guardian_phone=request.POST.get('guardian_phone', ''),
                        guardian_email=request.POST.get('guardian_email', '')
                    )
                    print(f"Created student profile")  # Debug log
                except Exception as e:
                    print(f"Error creating student profile: {str(e)}")  # Debug log
                    # Clean up the created user if profile creation fails
                    user.delete()
                    messages.error(request, f'Error creating student profile: {str(e)}')
                    return render(request, 'accounts/register.html')
                
                messages.success(request, 'Student account created successfully!')
                
            elif role == 'teacher':
                # Log teacher registration data
                print(f"Teacher registration data: {request.POST}")  # Debug log
                
                # Validate required fields
                required_fields = ['first_name', 'last_name', 'email', 'password1', 'password2', 'department', 'subjects', 'teacher_id']
                for field in required_fields:
                    if not request.POST.get(field):
                        print(f"Missing required field: {field}")  # Debug log
                        messages.error(request, f'{field.replace("_", " ").title()} is required.')
                        return render(request, 'accounts/register.html')
                
                # Check if email already exists
                email = request.POST.get('email')
                if User.objects.filter(email=email).exists():
                    print(f"Email already exists: {email}")  # Debug log
                    messages.error(request, 'Email already exists.')
                    return render(request, 'accounts/register.html')
                
                # Check if username (email) already exists
                if User.objects.filter(username=email).exists():
                    print(f"Username already exists: {email}")  # Debug log
                    messages.error(request, 'This email is already registered as a username.')
                    return render(request, 'accounts/register.html')
                
                # Validate password
                if request.POST.get('password1') != request.POST.get('password2'):
                    print("Passwords do not match")  # Debug log
                    messages.error(request, 'Passwords do not match.')
                    return render(request, 'accounts/register.html')
                
                try:
                    # Create user account for teacher
                    user = User.objects.create_user(
                        username=email,  # Using email as username for teachers
                        password=request.POST.get('password1'),
                        email=email,
                        first_name=request.POST.get('first_name'),
                        last_name=request.POST.get('last_name')
                    )
                    print(f"Created user account for teacher: {user.username}")  # Debug log
                except Exception as e:
                    print(f"Error creating user account: {str(e)}")  # Debug log
                    messages.error(request, f'Error creating user account: {str(e)}')
                    return render(request, 'accounts/register.html')
                
                try:
                    # Add user to teacher group
                    teacher_group = Group.objects.get_or_create(name='Teacher')[0]
                    user.groups.add(teacher_group)
                    print(f"Added user to teacher group")  # Debug log
                except Exception as e:
                    print(f"Error adding user to group: {str(e)}")  # Debug log
                    messages.error(request, f'Error adding user to group: {str(e)}')
                    return render(request, 'accounts/register.html')
                
                try:
                    # Create teacher profile with all required fields
                    Teacher.objects.create(
                        user=user,
                        teacher_id=request.POST.get('teacher_id'),  # Use provided teacher ID
                        first_name=request.POST.get('first_name'),
                        last_name=request.POST.get('last_name'),
                        email=email,
                        department=request.POST.get('department'),
                        subjects=request.POST.get('subjects', ''),
                        joining_date=datetime.now().date(),  # Using current date as joining date
                        phone_number=request.POST.get('phone_number', ''),
                        address=request.POST.get('address', '')
                    )
                    print(f"Created teacher profile")  # Debug log
                except Exception as e:
                    print(f"Error creating teacher profile: {str(e)}")  # Debug log
                    # Clean up the created user if profile creation fails
                    user.delete()
                    messages.error(request, f'Error creating teacher profile: {str(e)}')
                    return render(request, 'accounts/register.html')
                
                messages.success(request, 'Teacher account created successfully!')
                
            elif role == 'admin':
                # Log admin registration data
                print(f"Admin registration data: {request.POST}")  # Debug log
                
                # Validate required fields
                required_fields = ['username', 'email', 'password1', 'password2']
                for field in required_fields:
                    if not request.POST.get(field):
                        print(f"Missing required field: {field}")  # Debug log
                        messages.error(request, f'{field.replace("_", " ").title()} is required.')
                        return render(request, 'accounts/register.html')
                
                # Check if username already exists
                username = request.POST.get('username')
                if User.objects.filter(username=username).exists():
                    print(f"Username already exists: {username}")  # Debug log
                    messages.error(request, 'Username already exists.')
                    return render(request, 'accounts/register.html')
                
                # Check if email already exists
                email = request.POST.get('email')
                if User.objects.filter(email=email).exists():
                    print(f"Email already exists: {email}")  # Debug log
                    messages.error(request, 'Email already exists.')
                    return render(request, 'accounts/register.html')
                
                # Validate password
                if request.POST.get('password1') != request.POST.get('password2'):
                    print("Passwords do not match")  # Debug log
                    messages.error(request, 'Passwords do not match.')
                    return render(request, 'accounts/register.html')
                
                try:
                    # Create admin user account
                    user = User.objects.create_user(
                        username=username,
                        password=request.POST.get('password1'),
                        email=email,
                        is_staff=True,
                        is_superuser=True
                    )
                    print(f"Created admin user account: {user.username}")  # Debug log
                except Exception as e:
                    print(f"Error creating admin account: {str(e)}")  # Debug log
                    messages.error(request, f'Error creating admin account: {str(e)}')
                    return render(request, 'accounts/register.html')
                
                try:
                    # Add user to admin group (if it exists)
                    admin_group, created = Group.objects.get_or_create(name='Admin')
                    user.groups.add(admin_group)
                    print(f"Added user to admin group")  # Debug log
                except Exception as e:
                    print(f"Error adding user to admin group: {str(e)}")  # Debug log
                    # This is not critical, so we don't return an error
                
                messages.success(request, 'Admin account created successfully!')
                
            else:
                print(f"Invalid role specified: {role}")  # Debug log
                messages.error(request, 'Invalid role specified!')
                return redirect('register')
            
            return redirect('login')
            
        except ValidationError as e:
            print(f"Validation error during registration: {str(e)}")  # Debug log
            messages.error(request, str(e))
        except Exception as e:
            print(f"Unexpected error during registration: {str(e)}")  # Debug log
            messages.error(request, f'An error occurred during registration: {str(e)}')
            
    return render(request, 'accounts/register.html')

@login_required
def index_view(request):
    # Redirect to the proper dashboard view to ensure data is loaded
    return redirect('dashboard')

def logout_view(request):
    user = request.user
    role = "User"
    if user.is_authenticated:
        if hasattr(user, 'student_profile'):
            role = "Student"
        elif hasattr(user, 'teacher_profile'):
            role = "Teacher"
        elif user.is_staff:
            role = "Admin"
    logout(request)
    messages.success(request, f"{role} successfully logged out.")
    return redirect('login')

def forgot_password(request):
    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        
        if not email:
            messages.error(request, 'Please enter your email address.')
            return render(request, 'accounts/forgot_password.html')
            
        try:
            user = User.objects.get(email=email)
            
            # Generate a 6-digit verification code
            verification_code = ''.join([str(random.randint(0, 9)) for _ in range(6)])
            
            # Store the verification code and timestamp in session
            request.session['reset_code'] = verification_code
            request.session['reset_email'] = email
            request.session['code_timestamp'] = datetime.now().timestamp()
            request.session['code_attempts'] = 0  # Track verification attempts
            
            # Send email with verification code
            subject = 'Password Reset Verification Code - School Management System'
            message = f'''
            Hello {user.username},
            
            You have requested to reset your password. Your verification code is:
            
            {verification_code}
            
            This code will expire in 10 minutes for security purposes.
            If you did not request this password reset, please ignore this email.
            
            Best regards,
            School Management System Team
            '''
            
            from_email = settings.EMAIL_HOST_USER
            recipient_list = [email]
            
            try:
                send_mail(subject, message, from_email, recipient_list)
                messages.success(request, 'Verification code has been sent to your email. Please check your inbox.')
                return redirect('verify_code')
            except Exception as e:
                messages.error(request, 'Failed to send verification code. Please try again later.')
                return render(request, 'accounts/forgot_password.html')
                
        except User.DoesNotExist:
            # Use a generic message for security
            messages.error(request, 'If an account exists with this email, you will receive a verification code.')
            time.sleep(1)  # Add delay to prevent email enumeration attacks
    
    return render(request, 'accounts/forgot_password.html')

def verify_code(request):
    if 'reset_code' not in request.session or 'reset_email' not in request.session:
        messages.error(request, 'Please submit your email first.')
        return redirect('forgot_password')
        
    # Check if code has expired (10 minutes)
    timestamp = request.session.get('code_timestamp', 0)
    if (datetime.now().timestamp() - timestamp) > 600:
        messages.error(request, 'Verification code has expired. Please request a new one.')
        return redirect('forgot_password')
        
    if request.method == 'POST':
        entered_code = request.POST.get('verification_code', '').strip()
        stored_code = request.session.get('reset_code')
        attempts = request.session.get('code_attempts', 0)
        
        # Limit verification attempts
        if attempts >= 3:
            messages.error(request, 'Too many incorrect attempts. Please request a new code.')
            return redirect('forgot_password')
            
        if not entered_code:
            messages.error(request, 'Please enter the verification code.')
        elif entered_code == stored_code:
            request.session['code_verified'] = True
            messages.success(request, 'Code verified successfully.')
            return redirect('reset_password')
        else:
            request.session['code_attempts'] = attempts + 1
            remaining_attempts = 3 - (attempts + 1)
            messages.error(request, f'Invalid verification code. {remaining_attempts} attempts remaining.')
    
    return render(request, 'accounts/verify_code.html')

def reset_password(request):
    if not request.session.get('code_verified'):
        messages.error(request, 'Please verify your email first.')
        return redirect('forgot_password')
        
    if request.method == 'POST':
        password1 = request.POST.get('password1', '').strip()
        password2 = request.POST.get('password2', '').strip()
        
        # Validate password
        if not password1 or not password2:
            messages.error(request, 'Please enter both passwords.')
        elif password1 != password2:
            messages.error(request, 'Passwords do not match.')
        elif len(password1) < 8:
            messages.error(request, 'Password must be at least 8 characters long.')
        elif not any(c.isupper() for c in password1):
            messages.error(request, 'Password must contain at least one uppercase letter.')
        elif not any(c.islower() for c in password1):
            messages.error(request, 'Password must contain at least one lowercase letter.')
        elif not any(c.isdigit() for c in password1):
            messages.error(request, 'Password must contain at least one number.')
        elif not any(c in '!@#$%^&*()_+-=[]{}|;:,.<>?' for c in password1):
            messages.error(request, 'Password must contain at least one special character.')
        else:
            try:
                user = User.objects.get(email=request.session['reset_email'])
                user.set_password(password1)
                user.save()
                
                # Clear all session data
                for key in ['reset_code', 'reset_email', 'code_timestamp', 'code_attempts', 'code_verified']:
                    request.session.pop(key, None)
                
                messages.success(request, 'Your password has been reset successfully. You can now login with your new password.')
                return redirect('login')
                
            except User.DoesNotExist:
                messages.error(request, 'An error occurred. Please try again.')
                return redirect('forgot_password')
    
    return render(request, 'accounts/reset_password.html')

@never_cache
@login_required
def dashboard_view(request):
    from dateutil.relativedelta import relativedelta
    
    # Get current counts
    total_students = Student.objects.filter(status='active').count()
    total_teachers = Teacher.objects.filter(is_active=True).count()
    active_courses = Course.objects.count()
    
    # Calculate attendance statistics
    total_attendance_records = Attendance.objects.count()
    present_records = Attendance.objects.filter(status='present').count()
    absent_records = Attendance.objects.filter(status='absent').count()
    late_records = Attendance.objects.filter(status='late').count()
    
    attendance_rate = round((present_records / total_attendance_records * 100), 1) if total_attendance_records > 0 else 0
    
    # Calculate monthly growth (comparing with last month)
    today = timezone.now().date()
    first_day_this_month = today.replace(day=1)
    first_day_last_month = (first_day_this_month - relativedelta(months=1))
    
    # Student growth
    students_last_month = Student.objects.filter(
        created_at__date__lt=first_day_this_month,
        status='active'
    ).count()
    student_growth = round(((total_students - students_last_month) / students_last_month * 100), 1) if students_last_month > 0 else 0
    
    # Teacher growth
    teachers_last_month = Teacher.objects.filter(
        created_at__date__lt=first_day_this_month,
        is_active=True
    ).count()
    teacher_growth = round(((total_teachers - teachers_last_month) / teachers_last_month * 100), 1) if teachers_last_month > 0 else 0
    
    # Course changes
    courses_last_month = Course.objects.filter(
        created_at__date__lt=first_day_this_month
    ).count()
    course_change = round(((active_courses - courses_last_month) / courses_last_month * 100), 1) if courses_last_month > 0 else 0
    
    # Calculate previous month attendance for comparison
    previous_attendance = Attendance.objects.filter(
        date__year=first_day_last_month.year,
        date__month=first_day_last_month.month
    )
    prev_total = previous_attendance.count()
    prev_present = previous_attendance.filter(status='present').count()
    prev_attendance_rate = round((prev_present / prev_total * 100), 1) if prev_total > 0 else 0
    attendance_growth = round((attendance_rate - prev_attendance_rate), 1)
    
    # Student performance data (last 6 months)
    performance_data = []
    for i in range(5, -1, -1):  # Start from 5 months ago to current month
        month_date = today - relativedelta(months=i)
        month_name = month_date.strftime('%b')
        
        # Get average grades for this month
        month_grades = Grade.objects.filter(
            date__year=month_date.year,
            date__month=month_date.month
        )
        avg_score = month_grades.aggregate(avg_score=Avg('score'))['avg_score']
        
        # Use a baseline score if no grades found
        if avg_score is None:
            # Generate some realistic sample data based on month
            base_scores = [75, 78, 82, 85, 80, 88]
            avg_score = base_scores[i % len(base_scores)]
        
        performance_data.append({
            'month': month_name,
            'score': round(float(avg_score), 1)
        })
    
    # Attendance distribution
    attendance_distribution = {
        'present': round((present_records / total_attendance_records * 100), 1) if total_attendance_records > 0 else 85,
        'absent': round((absent_records / total_attendance_records * 100), 1) if total_attendance_records > 0 else 10,
        'late': round((late_records / total_attendance_records * 100), 1) if total_attendance_records > 0 else 5
    }
    
    context = {
        # Main statistics
        'total_students': total_students,
        'total_teachers': total_teachers,
        'active_courses': active_courses,
        'attendance_rate': attendance_rate,
        
        # Growth percentages
        'student_growth': student_growth,
        'teacher_growth': teacher_growth,
        'course_change': course_change,
        'attendance_growth': attendance_growth,
        
        # Chart data
        'performance_data': performance_data,
        'attendance_distribution': attendance_distribution,
        
        # Additional dashboard data
        'upcoming_events': Event.objects.filter(start_date__gte=timezone.now()).count(),
        'recent_activities': Activity.objects.all().order_by('-created_at')[:5],
        'recent_payments': Payment.objects.filter(status='completed').order_by('-payment_date')[:5] if Payment.objects.exists() else [],
        'pending_fees': Invoice.objects.filter(paid=False).count(),
    }
    

    response = render(request, 'accounts/dashboard.html', context)
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response

# Student Management Views
@login_required
def students_view(request):
    students = Student.objects.all()
    grades = Student.objects.values_list('grade_level', flat=True).distinct()
    context = {
        'students': students,
        'grades': sorted(grades),
        'statuses': Student.STATUS_CHOICES,
    }
    return render(request, 'accounts/students.html', context)

@login_required
def add_student_view(request):
    if request.method == 'POST':
        form = StudentRegistrationForm(request.POST, request.FILES)
        if form.is_valid():
            user = form.save()
            # Assuming Student is linked via user.student_profile or similar
            try:
                student = user.student_profile
            except AttributeError:
                # fallback if not using OneToOneField named student_profile
                from .models import Student
                student = Student.objects.get(user=user)
            messages.success(request, 'Student added successfully.')
            return redirect('student_detail', id=student.id)
    else:
        form = StudentRegistrationForm()
    context = {'form': form}
    return render(request, 'accounts/students/add_student.html', context)

@login_required
def student_detail_view(request, id):
    student = get_object_or_404(Student, id=id)
    context = {'student': student}
    return render(request, 'accounts/students/student_detail.html', context)

@login_required
def edit_student_view(request, id):
    try:
        print(f"Debug: edit_student_view called with id={id}")  # Debug log
        student = get_object_or_404(Student, id=id)
        user = student.user
        if request.method == 'POST':
            print(f"Debug: Processing POST request")  # Debug log
            form = StudentForm(request.POST, request.FILES, instance=student)
            password_form = StudentPasswordChangeForm(request.POST)
            if form.is_valid() and password_form.is_valid():
                form.save()
                password1 = password_form.cleaned_data.get('password1')
                if password1:
                    user.set_password(password1)
                    user.save()
                messages.success(request, 'Student information updated successfully.')
                return redirect('student_detail', id=student.id)
            else:
                print(f"Debug: Form errors: {form.errors}")  # Debug log
                messages.error(request, 'Please correct the errors below.')
        else:
            print(f"Debug: Processing GET request")  # Debug log
            form = StudentForm(instance=student)
            password_form = StudentPasswordChangeForm()
        context = {
            'form': form,
            'password_form': password_form,
            'student': student
        }
        return render(request, 'accounts/students/edit_student.html', context)
    except Exception as e:
        print(f"Debug: Error in edit_student_view: {str(e)}")  # Debug log
        messages.error(request, f'An error occurred: {str(e)}')
        return redirect('students')

@login_required
def delete_student_view(request, id):
    if not request.user.is_staff:
        messages.error(request, "You don't have permission to delete students.")
        return redirect('students')
        
    try:
        student = Student.objects.get(id=id)
        student.delete()
        messages.success(request, f'Student {student.first_name} {student.last_name} has been deleted successfully.')
    except Student.DoesNotExist:
        messages.error(request, 'Student not found.')
    
    return redirect('students')

# Teacher Management Views
@login_required
def teachers_view(request):
    teachers = Teacher.objects.all()
    departments = Teacher.objects.values_list('department', flat=True).distinct()
    context = {
        'teachers': teachers,
        'departments': sorted(departments),
    }
    return render(request, 'accounts/teachers.html', context)

@login_required
def add_teacher_view(request):
    if request.method == 'POST':
        form = TeacherRegistrationForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                user = form.save()
                teacher = user.teacher_profile
                messages.success(request, 'Teacher added successfully.')
                return redirect('teacher_detail', id=teacher.id)
            except Exception as e:
                messages.error(request, f'Error creating teacher: {str(e)}')
        else:
            # Display form errors
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = TeacherRegistrationForm()
    
    context = {'form': form}
    return render(request, 'accounts/teachers/add_teacher.html', context)

@login_required
def teacher_detail_view(request, id):
    teacher = get_object_or_404(Teacher, id=id)
    
    if request.method == 'POST':
        # Handle modal form submission for quick edit
        try:
            # Check for teacher_id uniqueness if it's being changed
            teacher_id = request.POST.get('teacher_id')
            if teacher_id and teacher_id != teacher.teacher_id:
                if Teacher.objects.filter(teacher_id=teacher_id).exclude(pk=teacher.pk).exists():
                    messages.error(request, 'A teacher with this ID already exists.')
                    return redirect('teacher_detail', id=id)
            
            # Update teacher fields
            teacher.teacher_id = teacher_id or teacher.teacher_id
            teacher.first_name = request.POST.get('first_name', teacher.first_name)
            teacher.last_name = request.POST.get('last_name', teacher.last_name)
            teacher.email = request.POST.get('email', teacher.email)
            teacher.phone_number = request.POST.get('phone_number', teacher.phone_number)
            teacher.department = request.POST.get('department', teacher.department)
            teacher.subjects = request.POST.get('subjects', teacher.subjects)
            teacher.qualification = request.POST.get('qualification', teacher.qualification)
            teacher.address = request.POST.get('address', teacher.address)
            teacher.is_active = request.POST.get('is_active') == 'True'
            
            # Handle date fields
            date_of_birth = request.POST.get('date_of_birth')
            if date_of_birth:
                teacher.date_of_birth = date_of_birth
                
            joining_date = request.POST.get('joining_date')
            if joining_date:
                teacher.joining_date = joining_date
            
            # Handle gender
            gender = request.POST.get('gender')
            if gender:
                teacher.gender = gender
            
            # Handle photo upload
            photo = request.FILES.get('photo')
            if photo:
                teacher.photo = photo
            
            teacher.save()
            
            # Update user email if needed
            if teacher.user and teacher.user.email != teacher.email:
                teacher.user.email = teacher.email
                teacher.user.save()
            
            messages.success(request, 'Teacher updated successfully.')
            return redirect('teacher_detail', id=id)
            
        except Exception as e:
            messages.error(request, f'Error updating teacher: {str(e)}')
            return redirect('teacher_detail', id=id)
    
    context = {'teacher': teacher}
    return render(request, 'accounts/teachers/teacher_detail.html', context)

@login_required
def edit_teacher_view(request, id):
    teacher = get_object_or_404(Teacher, id=id)
    
    if request.method == 'POST':
        form = TeacherForm(request.POST, request.FILES, instance=teacher)
        password_changed = False
        
        if form.is_valid():
            # Handle password change if provided
            current_password = request.POST.get('current_password')
            new_password1 = request.POST.get('new_password1')
            new_password2 = request.POST.get('new_password2')
            
            # Check if password change was requested
            if new_password1 or new_password2 or current_password:
                if not current_password:
                    messages.error(request, 'Current password is required to change password.')
                elif not new_password1:
                    messages.error(request, 'New password is required.')
                elif new_password1 != new_password2:
                    messages.error(request, 'New passwords do not match.')
                elif teacher.user and not check_password(current_password, teacher.user.password):
                    messages.error(request, 'Current password is incorrect.')
                else:
                    # Validate password strength
                    if len(new_password1) < 8:
                        messages.error(request, 'Password must be at least 8 characters long.')
                    elif not any(c.islower() for c in new_password1):
                        messages.error(request, 'Password must contain at least one lowercase letter.')
                    elif not any(c.isupper() for c in new_password1):
                        messages.error(request, 'Password must contain at least one uppercase letter.')
                    elif not any(c.isdigit() for c in new_password1):
                        messages.error(request, 'Password must contain at least one number.')
                    else:
                        # Password is valid, update it
                        teacher.user.set_password(new_password1)
                        teacher.user.save()
                        password_changed = True
            
            # Save the teacher form
            try:
                form.save()
                success_msg = 'Teacher updated successfully.'
                if password_changed:
                    success_msg += ' Password has been changed.'
                messages.success(request, success_msg)
                return redirect('teacher_detail', id=teacher.id)
            except Exception as e:
                messages.error(request, f'Error saving teacher: {str(e)}')
    else:
        form = TeacherForm(instance=teacher)
    
    context = {
        'form': form,
        'teacher': teacher
    }
    return render(request, 'accounts/teachers/edit_teacher.html', context)

@login_required
def delete_teacher_view(request, id):
    teacher = get_object_or_404(Teacher, id=id)
    
    # Check if user has permission to delete the teacher
    if not request.user.is_staff:
        messages.error(request, 'You do not have permission to delete teachers.')
        return redirect('teacher_detail', id=id)
    
    if request.method == 'POST':
        try:
            # Get the teacher's name before deletion for the success message
            teacher_name = teacher.get_full_name()
            
            # Delete the teacher
            teacher.delete()
            
            messages.success(request, f'Teacher "{teacher_name}" deleted successfully.')
            return redirect('teachers')
        except Exception as e:
            messages.error(request, f'Error deleting teacher: {str(e)}')
            return redirect('teacher_detail', id=id)
    
    context = {
        'teacher': teacher
    }
    return render(request, 'accounts/teachers/delete_teacher.html', context)

# Course Management Views
@login_required
def courses_view(request):
    courses = Course.objects.all()
    credits = Course.objects.values_list('credits', flat=True).distinct()
    context = {
        'courses': courses,
        'credits': sorted(credits),
    }
    return render(request, 'accounts/courses.html', context)

@login_required
def add_course_view(request):
    if request.method == 'POST':
        form = CourseForm(request.POST)
        if form.is_valid():
            course = form.save()
            messages.success(request, 'Course added successfully.')
            return redirect('course_detail', id=course.id)
    else:
        form = CourseForm()
    
    context = {'form': form}
    return render(request, 'accounts/courses/add_course.html', context)

@login_required
def course_detail_view(request, id):
    course = get_object_or_404(Course, id=id)
    context = {'course': course}
    return render(request, 'accounts/courses/course_detail.html', context)

@login_required
def edit_course_view(request, id):
    course = get_object_or_404(Course, id=id)
    if request.method == 'POST':
        form = CourseForm(request.POST, instance=course)
        if form.is_valid():
            form.save()
            messages.success(request, 'Course updated successfully.')
            return redirect('course_detail', id=course.id)
    else:
        form = CourseForm(instance=course)
    
    context = {
        'form': form,
        'course': course
    }
    return render(request, 'accounts/courses/edit_course.html', context)

@login_required
def delete_course_view(request, id):
    course = get_object_or_404(Course, id=id)
    
    # Check if user has permission to delete the course
    if not request.user.is_staff and (not hasattr(request.user, 'teacher_profile') or request.user.teacher_profile != course.teacher):
        messages.error(request, 'You do not have permission to delete this course.')
        return redirect('course_detail', id=id)
    
    if request.method == 'POST':
        try:
            course.delete()
            messages.success(request, 'Course deleted successfully.')
            return redirect('courses')
        except Exception as e:
            messages.error(request, f'Error deleting course: {str(e)}')
            return redirect('course_detail', id=id)
    
    context = {
        'course': course
    }
    return render(request, 'accounts/courses/delete_course.html', context)

# Finance Management Views
@login_required
def fee_payment_view(request):
    # Get search query
    search_query = request.GET.get('search', '')
    
    # Get all payments with search filter
    payments = Payment.objects.all()
    
    if search_query:
        payments = payments.filter(
            Q(transaction_id__icontains=search_query) |
            Q(student__first_name__icontains=search_query) |
            Q(student__last_name__icontains=search_query) |
            Q(payment_method__icontains=search_query) |
            Q(status__icontains=search_query) |
            Q(amount__icontains=search_query)
        )
    
    payments = payments.order_by('-payment_date')  # Order by payment date, most recent first
    
    # Calculate payment statistics
    total_collected = payments.aggregate(total=models.Sum('amount'))['total'] or 0
    paid_amount = payments.filter(status='completed').aggregate(total=models.Sum('amount'))['total'] or 0
    pending_amount = payments.filter(status='pending').aggregate(total=models.Sum('amount'))['total'] or 0
    overdue_amount = payments.filter(status='overdue').aggregate(total=models.Sum('amount'))['total'] or 0
    
    # Handle payment creation
    if request.method == 'POST':
        form = PaymentForm(request.POST)
        if form.is_valid():
            payment = form.save(commit=False)
            payment.status = 'completed'  # Set initial status
            payment.transaction_id = f"TXN{int(time.time())}"  # Generate unique transaction ID
            payment.save()
            messages.success(request, 'Payment added successfully.')
            return redirect('fee_payment')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = PaymentForm()
    
    context = {
        'page_title': 'Fee Payment',
        'payments': payments,
        'total_collected': total_collected,
        'paid_amount': paid_amount,
        'pending_amount': pending_amount,
        'overdue_amount': overdue_amount,
        'form': form,
        'students': Student.objects.all(),  # Add students for the dropdown
        'search_query': search_query  # Pass search query back to template
    }
    return render(request, 'accounts/fee_payment.html', context)

@login_required
def invoices_view(request):
    # Get all invoices
    invoices = Invoice.objects.all().order_by('-created_at')
    
    # Calculate invoice statistics
    total_amount = invoices.aggregate(total=models.Sum('amount'))['total'] or 0
    paid_amount = invoices.filter(paid=True).aggregate(total=models.Sum('amount'))['total'] or 0
    pending_amount = invoices.filter(paid=False).aggregate(total=models.Sum('amount'))['total'] or 0
    overdue_amount = invoices.filter(paid=False, due_date__lt=datetime.now().date()).aggregate(total=models.Sum('amount'))['total'] or 0
    
    # Handle invoice creation
    if request.method == 'POST':
        form = InvoiceForm(request.POST)
        if form.is_valid():
            invoice = form.save(commit=False)
            invoice.created_at = datetime.now()
            invoice.paid = False
            invoice.save()
            messages.success(request, 'Invoice created successfully.')
            return redirect('invoices')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = InvoiceForm()
    
    context = {
        'page_title': 'Invoices',
        'invoices': invoices,
        'total_amount': total_amount,
        'paid_amount': paid_amount,
        'pending_amount': pending_amount,
        'overdue_amount': overdue_amount,
        'form': form
    }
    return render(request, 'accounts/invoices.html', context)

@login_required
def invoice_detail_view(request, id):
    invoice = get_object_or_404(Invoice, id=id)
    context = {'invoice': invoice}
    return render(request, 'accounts/finance/invoice_detail.html', context)

# Schedule Management Views
@login_required
def calendar_view(request):
    if request.method == 'POST':
        form = EventForm(request.POST)
        if form.is_valid():
            event = form.save(commit=False)
            event.created_by = request.user
            event.save()
            messages.success(request, 'Event created successfully.')
            return JsonResponse({'success': True, 'message': 'Event created successfully'})
        else:
            return JsonResponse({'success': False, 'errors': form.errors})
    
    # Get upcoming events for sidebar
    upcoming_events = Event.objects.filter(
        start_date__gte=datetime.now()
    ).order_by('start_date')[:5]
    
    # If this is a request for upcoming events only (for AJAX refresh)
    if request.GET.get('upcoming_events_only'):
        context = {
            'upcoming_events': upcoming_events,
        }
        return render(request, 'accounts/calendar.html', context)
    
    context = {
        'page_title': 'Calendar',
        'upcoming_events': upcoming_events,
        'form': EventForm()
    }
    return render(request, 'accounts/calendar.html', context)

@login_required
def calendar_events_api(request):
    """API endpoint to provide events for FullCalendar"""
    from datetime import date, timedelta
    import calendar as cal
    
    events = Event.objects.all()
    
    event_list = []
    
    # Add regular events from database
    for event in events:
        event_data = {
            'id': event.id,
            'title': event.title,
            'start': event.start_date.isoformat(),
            'backgroundColor': event.color,
            'borderColor': event.color,
            'textColor': '#ffffff' if event.color in ['#e74c3c', '#9b59b6', '#34495e'] else '#000000',
            'className': f'event-type-{event.event_type}',
            'extendedProps': {
                'description': event.description,
                'location': event.location,
                'event_type': event.get_event_type_display(),
                'created_by': event.created_by.get_full_name() if event.created_by else 'System',
                'is_database_event': True
            }
        }
        
        if event.end_date:
            event_data['end'] = event.end_date.isoformat()
            
        if event.is_all_day:
            event_data['allDay'] = True
        
        event_list.append(event_data)
    
    # Add Sunday holidays automatically
    today = date.today()
    start_date = today.replace(day=1) - timedelta(days=30)  # Start from last month
    end_date = today + timedelta(days=90)  # Go 3 months ahead
    
    current_date = start_date
    sunday_counter = 1
    
    while current_date <= end_date:
        if current_date.weekday() == 6:  # Sunday is 6 in Python's weekday()
            # Check if there's already a holiday event on this Sunday
            existing_sunday_event = any(
                event for event in events 
                if event.start_date.date() == current_date and event.event_type == 'holiday'
            )
            
            if not existing_sunday_event:
                sunday_event = {
                    'id': f'sunday-{current_date.isoformat()}',
                    'title': 'Sunday Holiday',
                    'start': current_date.isoformat(),
                    'backgroundColor': '#2ecc71',  # Holiday green color
                    'borderColor': '#27ae60',
                    'textColor': '#ffffff',
                    'className': 'event-type-holiday sunday-holiday',
                    'allDay': True,
                    'extendedProps': {
                        'description': 'Weekly Sunday holiday - a day of rest',
                        'location': '',
                        'event_type': 'Holiday',
                        'created_by': 'System',
                        'is_database_event': False,
                        'is_sunday_holiday': True
                    }
                }
                event_list.append(sunday_event)
        
        current_date += timedelta(days=1)
    
    return JsonResponse(event_list, safe=False)

@login_required
def event_detail_api(request, event_id):
    """API endpoint to get/delete event details"""
    
    if request.method == 'DELETE':
        # Handle event deletion
        if str(event_id).startswith('sunday-'):
            return JsonResponse({'error': 'Sunday holidays cannot be deleted'}, status=400)
        
        try:
            event = Event.objects.get(id=event_id)
            event.delete()
            return JsonResponse({'success': True, 'message': 'Event deleted successfully'})
        except Event.DoesNotExist:
            return JsonResponse({'error': 'Event not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    # Handle GET request for event details
    # Handle Sunday holiday events
    if str(event_id).startswith('sunday-'):
        date_str = event_id.replace('sunday-', '')
        try:
            from datetime import datetime
            sunday_date = datetime.fromisoformat(date_str)
            data = {
                'id': event_id,
                'title': 'Sunday Holiday',
                'description': 'Weekly Sunday holiday - a day of rest',
                'start_date': sunday_date.strftime('%Y-%m-%d'),
                'end_date': '',
                'location': '',
                'event_type': 'Holiday',
                'color': '#2ecc71',
                'is_all_day': True,
                'created_by': 'System',
                'created_at': 'System Generated'
            }
            return JsonResponse(data)
        except:
            return JsonResponse({'error': 'Invalid Sunday holiday event'}, status=404)
    
    # Handle regular database events
    try:
        event = Event.objects.get(id=event_id)
        data = {
            'id': event.id,
            'title': event.title,
            'description': event.description,
            'start_date': event.start_date.strftime('%Y-%m-%d %H:%M'),
            'end_date': event.end_date.strftime('%Y-%m-%d %H:%M') if event.end_date else '',
            'location': event.location,
            'event_type': event.get_event_type_display(),
            'color': event.color,
            'is_all_day': event.is_all_day,
            'created_by': event.created_by.get_full_name() if event.created_by else 'System',
            'created_at': event.created_at.strftime('%Y-%m-%d %H:%M')
        }
        return JsonResponse(data)
    except Event.DoesNotExist:
        return JsonResponse({'error': 'Event not found'}, status=404)

@login_required
def timetable_view(request):
    if request.method == 'POST':
        print("POST request received in timetable_view")
        print("POST data:", request.POST)
        
        form = ScheduleForm(request.POST)
        print("Form is valid:", form.is_valid())
        if not form.is_valid():
            print("Form errors:", form.errors)
        
        if form.is_valid():
            try:
                # Get the course and its class section
                course = form.cleaned_data['course']
                class_section = course.class_section
                
                # Check for schedule conflicts within the same class section
                conflicting_schedules = Schedule.objects.filter(
                    course__class_section=class_section,
                    day=form.cleaned_data['day']
                ).exclude(
                    course=course
                )
                
                # Check time overlap
                start_time = form.cleaned_data['start_time']
                end_time = form.cleaned_data['end_time']
                
                for schedule in conflicting_schedules:
                    if (start_time < schedule.end_time and 
                        end_time > schedule.start_time):
                        return JsonResponse({
                            'success': False,
                            'errors': {
                                'general': [
                                    f'Time slot conflicts with {schedule.course.title} '
                                    f'({schedule.start_time.strftime("%H:%M")} - {schedule.end_time.strftime("%H:%M")})'
                                ]
                            }
                        })
                
                schedule = form.save()
                
                # Get the updated schedule data for immediate display
                schedule_data = {
                    'id': schedule.id,
                    'course_title': schedule.course.title,
                    'teacher_name': schedule.course.teacher.get_full_name(),
                    'room': schedule.room,
                    'day': schedule.day,
                    'start_time': schedule.start_time.strftime('%H:%M'),
                    'end_time': schedule.end_time.strftime('%H:%M'),
                    'display_time': f"{schedule.start_time.strftime('%H:%M')} - {schedule.end_time.strftime('%H:%M')}"
                }
                
                messages.success(request, f'Class "{schedule.course.title}" has been added to the timetable successfully.')
                return JsonResponse({
                    'success': True, 
                    'message': f'Class "{schedule.course.title}" has been added successfully!',
                    'schedule': schedule_data
                })
            except Exception as e:
                print(f"Error saving schedule: {str(e)}")
                return JsonResponse({
                    'success': False, 
                    'errors': {'general': [f'Error saving schedule: {str(e)}']}
                })
        else:
            print("Form validation failed")
            return JsonResponse({'success': False, 'errors': form.errors})
    
    # Get selected class section from query parameters
    selected_class = request.GET.get('class')
    
    # Get all schedules organized by day and time
    schedules = Schedule.objects.select_related('course', 'course__teacher', 'course__class_section')
    
    # Filter schedules by class section if selected
    if selected_class:
        schedules = schedules.filter(course__class_section_id=selected_class)
    
    schedules = schedules.order_by('day', 'start_time')
    
    # Get all available courses and classes for filters
    courses = Course.objects.all()
    if selected_class:
        courses = courses.filter(class_section_id=selected_class)
    
    # Serialize courses for JavaScript
    courses_data = [
        {
            'id': course.id,
            'title': course.title,
            'class_section': course.class_section_id,
            'course_code': course.course_code,
            'teacher_name': course.teacher.get_full_name() if course.teacher else ''
        } for course in courses
    ]
    
    classes = Class.objects.all()
    
    # Create form instance with filtered courses
    form = ScheduleForm()
    form.fields['course'].queryset = courses
    
    # Get custom time slots from request or use defaults
    start_hour = int(request.GET.get('start_hour', 8))  # Default 8 AM
    end_hour = int(request.GET.get('end_hour', 17))     # Default 5 PM
    slot_duration = int(request.GET.get('duration', 60))  # Default 60 minutes
    
    # Create time slots based on settings
    time_slots = []
    current_hour = start_hour
    current_minute = 0
    
    while current_hour < end_hour:
        start_time = f"{current_hour:02d}:{current_minute:02d}"
        
        # Calculate end time
        end_minute = current_minute + slot_duration
        end_hour_calc = current_hour
        
        if end_minute >= 60:
            end_hour_calc += end_minute // 60
            end_minute = end_minute % 60
            
        end_time = f"{end_hour_calc:02d}:{end_minute:02d}"
        
        time_slots.append({
            'start': start_time,
            'end': end_time,
            'display': f"{start_time} - {end_time}"
        })
        
        # Move to next slot
        current_minute += slot_duration
        if current_minute >= 60:
            current_hour += current_minute // 60
            current_minute = current_minute % 60
    
    # Organize schedules by day and time for easier template rendering
    days = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday']
    timetable_data = []
    
    for time_slot in time_slots:
        row_data = {
            'time_slot': time_slot,
            'cells': []
        }
        
        for day in days:
            # Find schedule for this day and time
            from datetime import time
            time_start = time.fromisoformat(time_slot['start'])
            schedule = schedules.filter(
                day=day,
                start_time__lte=time_start,
                end_time__gt=time_start
            ).first()
            
            row_data['cells'].append({
                'day': day,
                'schedule': schedule,
                'occupied': schedule is not None
            })
        
        timetable_data.append(row_data)
    
    context = {
        'page_title': 'Timetable',
        'form': form,
        'schedules': schedules,
        'courses': courses,
        'courses_json': courses_data,
        'classes': classes,
        'time_slots': time_slots,
        'timetable_data': timetable_data,
        'days': days,
        'start_hour': start_hour,
        'end_hour': end_hour,
        'slot_duration': slot_duration,
        'selected_class': selected_class
    }
    
    # If it's an AJAX request, return only the timetable portion
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'accounts/timetable_partial.html', context)
    
    return render(request, 'accounts/timetable.html', context)

@login_required
def timetable_print(request):
    # Get selected class section from query parameters
    selected_class = request.GET.get('class')
    
    # Get all schedules organized by day and time
    schedules = Schedule.objects.select_related('course', 'course__teacher', 'course__class_section')
    
    # Filter schedules by class section if selected
    if selected_class:
        schedules = schedules.filter(course__class_section_id=selected_class)
        selected_class_obj = Class.objects.get(id=selected_class)
        selected_class_name = f"{selected_class_obj.name} - {selected_class_obj.section}"
    else:
        selected_class_name = "All Classes"

    # Create schedule dictionary for easy lookup
    schedule_dict = {}
    for schedule in schedules:
        time_key = schedule.start_time.strftime('%H:%M')
        if time_key not in schedule_dict:
            schedule_dict[time_key] = {}
        schedule_dict[time_key][schedule.day] = schedule

    # Get time settings
    start_hour = int(request.GET.get('start_hour', 8))
    end_hour = int(request.GET.get('end_hour', 17))
    duration = int(request.GET.get('duration', 60))

    # Generate time slots
    time_slots = []
    for hour in range(start_hour, end_hour):
        for minute in range(0, 60, duration):
            time_slots.append(f"{hour:02d}:{minute:02d}")

    # Prepare context
    context = {
        'selected_class_name': selected_class_name,
        'start_hour': start_hour,
        'end_hour': end_hour,
        'slot_duration': duration,
        'time_slots': time_slots,
        'schedule_dict': schedule_dict,
        'days': ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday']
    }
    
    return render(request, 'accounts/timetable_print.html', context)

# Academic Management Views
@login_required
def reports_view(request):
    students = Student.objects.filter(status='active')
    teachers = Teacher.objects.filter(is_active=True)
    classes = Class.objects.all()
    departments = Department.objects.all()

    # --- Statistics ---
    total_students = students.count()

    # Attendance: % present out of all attendance records for active students
    attendance_records = Attendance.objects.filter(student__in=students)
    total_attendance = attendance_records.count()
    present_attendance = attendance_records.filter(status='present').count()
    average_attendance = round((present_attendance / total_attendance * 100), 2) if total_attendance > 0 else 0

    # Grades: average grade and pass rate
    grades = Grade.objects.filter(student__in=students)
    average_grade = round(grades.aggregate(avg=Avg('score'))['avg'] or 0, 2)
    pass_count = grades.filter(score__gte=60).count()  # Assuming 60 is pass mark
    total_grades = grades.count()
    pass_rate = round((pass_count / total_grades * 100), 2) if total_grades > 0 else 0

    # Grade Distribution for chart
    grade_distribution = []
    for letter in ['A', 'B', 'C', 'D', 'F']:
        count = int(grades.filter(grade=letter).count())
        grade_distribution.append({'grade': letter, 'count': count})
    grade_distribution_json = json.dumps(grade_distribution)

    # Subject Performance for chart (average score per subject)
    subject_performance = []
    subjects = Course.objects.all()
    for subject in subjects:
        subject_grades = grades.filter(subject=subject)
        if subject_grades.exists():
            avg_score = subject_grades.aggregate(avg=Avg('score'))['avg'] or 0
            avg_score = float(avg_score)
            subject_performance.append({'subject': subject.title, 'average_score': avg_score})
    subject_performance_json = json.dumps(subject_performance)

    context = {
        'page_title': 'Reports',
        'students': students,
        'teachers': teachers,
        'classes': classes,
        'departments': departments,
        'average_attendance': average_attendance,
        'average_grade': average_grade,
        'pass_rate': pass_rate,
        'total_students': total_students,
        'grade_distribution_json': grade_distribution_json,
        'subject_performance_json': subject_performance_json,
    }
    return render(request, 'accounts/reports.html', context)

@login_required
def examinations_view(request):
    # Get all examinations with their schedules
    examinations = Examination.objects.select_related('course', 'course__teacher').prefetch_related('examschedule_set').order_by('-date')
    
    # Process examinations to add status and formatting
    for exam in examinations:
        # Determine status based on date
        now = timezone.now()
        if exam.date < now:
            exam.status = 'completed'
        elif exam.date <= now + timedelta(days=7):  # Ongoing if within a week
            exam.status = 'upcoming'
        else:
            exam.status = 'upcoming'
            
        # Get schedule info
        exam_schedule = exam.examschedule_set.first()
        exam.room = exam_schedule.room if exam_schedule else 'Not assigned'
        exam.supervisor = exam_schedule.supervisor if exam_schedule else None
        
        # Format datetime
        exam.formatted_date = exam.date.strftime('%B %d, %Y at %I:%M %p')
        exam.formatted_duration = str(exam.duration)
    
    # Calculate statistics
    total_exams = examinations.count()
    upcoming_exams = examinations.filter(date__gt=timezone.now()).count()
    completed_exams = examinations.filter(date__lt=timezone.now()).count()
    
    # Get recent results summary (you can customize this based on your needs)
    pass_rate = 85  # Placeholder - calculate based on actual results
    highest_score = 92  # Placeholder
    lowest_score = 68  # Placeholder
    total_students = Student.objects.filter(status='active').count()
    
    context = {
        'page_title': 'Examinations',
        'examinations': examinations,
        'exams': examinations,  # For backward compatibility
        'total_exams': total_exams,
        'upcoming_exams': upcoming_exams,
        'completed_exams': completed_exams,
        'pass_rate': pass_rate,
        'highest_score': highest_score,
        'lowest_score': lowest_score,
        'total_students': total_students
    }
    return render(request, 'accounts/examinations.html', context)

@login_required
def exam_schedule_view(request):
    if request.method == 'POST':
        form = ExaminationForm(request.POST)
        schedule_form = ExamScheduleForm(request.POST)
        
        if form.is_valid() and schedule_form.is_valid():
            try:
                # Save the examination
                examination = form.save()
                
                # Save the exam schedule
                exam_schedule = schedule_form.save(commit=False)
                exam_schedule.examination = examination
                exam_schedule.save()
                
                # Create an event for the calendar
                Event.objects.create(
                    title=f"Exam: {examination.title or examination.exam_type}",
                    description=f"Course: {examination.course.title}\nDuration: {examination.duration}\nTotal Marks: {examination.total_marks}\nRoom: {exam_schedule.room}\nSupervisor: {exam_schedule.supervisor.get_full_name()}",
                    start_date=examination.date,
                    end_date=examination.date + examination.duration,
                    location=exam_schedule.room,
                    event_type='exam',
                    color='#e74c3c',
                    created_by=request.user
                )
                
                messages.success(request, f'Exam "{examination.title or examination.exam_type}" has been scheduled successfully!')
                return redirect('examinations')
                
            except Exception as e:
                messages.error(request, f'Error scheduling exam: {str(e)}')
        else:
            # Handle form errors
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
            for field, errors in schedule_form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = ExaminationForm()
        schedule_form = ExamScheduleForm()
    
    # Get all courses and classes for the form
    courses = Course.objects.filter(teacher__is_active=True).order_by('course_code')
    classes = Class.objects.all().order_by('name', 'section')
    teachers = Teacher.objects.filter(is_active=True).order_by('first_name', 'last_name')
    
    context = {
        'page_title': 'Schedule Exam',
        'form': form,
        'schedule_form': schedule_form,
        'courses': courses,
        'classes': classes,
        'teachers': teachers
    }
    return render(request, 'accounts/exam_schedule.html', context)

@login_required
def exam_detail_view(request, exam_id):
    """Admin view for exam details"""
    exam = get_object_or_404(Examination, id=exam_id)
    
    # Get exam schedule
    exam_schedule = exam.examschedule_set.first()
    
    # Get students enrolled in the course
    students = exam.course.students.all()
    
    # Get results if any
    results = {}
    exam_date = exam.date.date()
    
    # Only show results if the grade is linked to this examination
    for student in students:
        try:
            grade = Grade.objects.filter(
                student=student,
                subject=exam.course,
                examination=exam
            ).first()
            if grade:
                results[student.id] = grade.score
            else:
                results[student.id] = None
        except Exception:
            results[student.id] = None
    
    # Calculate status variables
    now = timezone.now()
    upcoming_date = now + timedelta(days=7)
    
    # Calculate submitted results count
    submitted_results_count = sum(1 for result in results.values() if result is not None)
    
    context = {
        'exam': exam,
        'exam_schedule': exam_schedule,
        'students': students,
        'results': results,
        'now': now,
        'upcoming_date': upcoming_date,
        'submitted_results_count': submitted_results_count,
        'page_title': f'Exam Details - {exam.title or exam.exam_type}'
    }
    return render(request, 'accounts/exam_detail.html', context)

@login_required
def exam_edit_view(request, exam_id):
    """Admin view for editing exams"""
    exam = get_object_or_404(Examination, id=exam_id)
    
    if request.method == 'POST':
        form = ExaminationForm(request.POST, instance=exam)
        schedule_form = ExamScheduleForm(request.POST)
        
        if form.is_valid():
            try:
                # Save the examination
                examination = form.save()
                
                # Update or create the exam schedule
                exam_schedule = exam.examschedule_set.first()
                if exam_schedule:
                    schedule_form = ExamScheduleForm(request.POST, instance=exam_schedule)
                else:
                    schedule_form = ExamScheduleForm(request.POST)
                
                if schedule_form.is_valid():
                    exam_schedule = schedule_form.save(commit=False)
                    exam_schedule.examination = examination
                    exam_schedule.save()
                
                messages.success(request, f'Exam "{examination.title or examination.exam_type}" has been updated successfully!')
                return redirect('examinations')
                
            except Exception as e:
                messages.error(request, f'Error updating exam: {str(e)}')
        else:
            # Handle form errors
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = ExaminationForm(instance=exam)
        exam_schedule = exam.examschedule_set.first()
        if exam_schedule:
            schedule_form = ExamScheduleForm(instance=exam_schedule)
        else:
            schedule_form = ExamScheduleForm()
    
    # Get all courses and classes for the form
    courses = Course.objects.filter(teacher__is_active=True).order_by('course_code')
    classes = Class.objects.all().order_by('name', 'section')
    teachers = Teacher.objects.filter(is_active=True).order_by('first_name', 'last_name')
    
    context = {
        'page_title': f'Edit Exam - {exam.title or exam.exam_type}',
        'form': form,
        'schedule_form': schedule_form,
        'courses': courses,
        'classes': classes,
        'teachers': teachers,
        'exam': exam
    }
    return render(request, 'accounts/exam_edit.html', context)

@login_required
def exam_delete_view(request, exam_id):
    """Admin view for deleting exams"""
    exam = get_object_or_404(Examination, id=exam_id)
    
    if request.method == 'POST':
        try:
            exam_title = exam.title or exam.exam_type
            exam.delete()
            messages.success(request, f'Exam "{exam_title}" has been deleted successfully!')
            return redirect('examinations')
        except Exception as e:
            messages.error(request, f'Error deleting exam: {str(e)}')
            return redirect('exam_detail', exam_id=exam.id)
    
    context = {
        'exam': exam,
        'page_title': f'Delete Exam - {exam.title or exam.exam_type}'
    }
    return render(request, 'accounts/exam_delete.html', context)

# Settings and Profile Views
@login_required
def settings_view(request):
    context = {
        'page_title': 'Settings'
    }
    return render(request, 'accounts/settings.html', context)

@login_required
def profile_view(request):
    if request.method == 'POST':
        # Handle profile information update
        if 'first_name' in request.POST:
            user = request.user
            user.first_name = request.POST.get('first_name')
            user.last_name = request.POST.get('last_name')
            user.email = request.POST.get('email')
            user.save()
            
            # Update or create profile
            profile, created = Profile.objects.get_or_create(user=user)
            profile.bio = request.POST.get('bio', '')
            profile.save()
            
            messages.success(request, 'Profile updated successfully.')
            return redirect('profile')
            
        # Handle avatar update
        elif request.FILES.get('avatar'):
            profile, created = Profile.objects.get_or_create(user=request.user)
            profile.avatar = request.FILES['avatar']
            profile.save()
            messages.success(request, 'Profile picture updated successfully.')
            return redirect('profile')
    
    context = {
        'page_title': 'Profile'
    }
    return render(request, 'accounts/profile.html', context)

@login_required
def change_password(request):
    if request.method == 'POST':
        current_password = request.POST.get('current_password')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')
        
        user = request.user
        
        # Validate current password
        if not user.check_password(current_password):
            messages.error(request, 'Current password is incorrect.')
            return redirect('profile')
            
        # Validate new password
        if new_password != confirm_password:
            messages.error(request, 'New passwords do not match.')
            return redirect('profile')
            
        # Validate password strength
        if len(new_password) < 8:
            messages.error(request, 'Password must be at least 8 characters long.')
        elif not any(c.isupper() for c in new_password):
            messages.error(request, 'Password must contain at least one uppercase letter.')
        elif not any(c.islower() for c in new_password):
            messages.error(request, 'Password must contain at least one lowercase letter.')
        elif not any(c.isdigit() for c in new_password):
            messages.error(request, 'Password must contain at least one number.')
        elif not any(c in '!@#$%^&*()_+-=[]{}|;:,.<>?' for c in new_password):
            messages.error(request, 'Password must contain at least one special character.')
        else:
            # Change password
            user.set_password(new_password)
            user.save()
            
            # Update session to prevent logout
            update_session_auth_hash(request, user)
            
            messages.success(request, 'Your password has been changed successfully.')
            return redirect('profile')
            
    return redirect('profile')

# Student Views
@login_required
def student_dashboard(request):
    try:
        print('DEBUG: Entered student_dashboard view')
        if not hasattr(request.user, 'student_profile'):
            messages.error(request, 'You do not have access to the student dashboard.')
            return redirect('login')
        student = request.user.student_profile
        current_grade = student.grade_level if hasattr(student, 'grade_level') else "Not Set"
        enrolled_courses = student.enrolled_courses.all()
        enrolled_courses_with_progress = []
        for course in enrolled_courses:
            try:
                course.progress = course.get_progress_for_student(student)
            except Exception as e:
                print(f"DEBUG: Error in get_progress_for_student: {e}")
                course.progress = 0
            enrolled_courses_with_progress.append(course)
        all_courses = Course.objects.all()
        available_courses = []
        for course in all_courses:
            try:
                if course.can_student_enroll(student):
                    available_courses.append(course)
            except Exception as e:
                print(f"DEBUG: Error in can_student_enroll: {e}")
        
        # Get combined notices and announcements
        notices = list(Notice.objects.all())
        announcements = list(Announcement.objects.filter(course__in=enrolled_courses, is_active=True))
        combined_notices = notices + announcements
        recent_notices = sorted(combined_notices, key=lambda x: x.created_at, reverse=True)[:3]
        
        context = {
            'student': student,
            'enrolled_courses': enrolled_courses_with_progress,
            'available_courses': available_courses[:5],
            'upcoming_assignments': Activity.objects.filter(
                course__in=enrolled_courses,
                due_date__gte=datetime.now()
            ).order_by('due_date')[:5],
            'upcoming_exams': Examination.objects.filter(
                course__in=enrolled_courses,
                date__gte=datetime.now()
            ).order_by('date')[:5],
            'recent_notices': recent_notices,
            'pending_fees': Invoice.objects.filter(
                student=student,
                paid=False
            ).first(),
            'current_grade': current_grade
        }
        # Attach schedule info to each upcoming exam
        for exam in context['upcoming_exams']:
            schedule = exam.examschedule_set.first()
            exam.start_time = schedule.start_time if schedule else None
            exam.venue = schedule.room if schedule else "Not assigned"
        print('DEBUG: student_dashboard context:', context)
        return render(request, 'accounts/student/dashboard.html', context)
    except Exception as e:
        print(f"DEBUG: Exception in student_dashboard: {e}")
        messages.error(request, f"An error occurred in the dashboard: {e}")
        return render(request, 'accounts/student/dashboard.html', {'student': request.user.student_profile, 'enrolled_courses': [], 'available_courses': [], 'upcoming_assignments': [], 'upcoming_exams': [], 'recent_notices': [], 'pending_fees': None, 'current_grade': 'Not Set'})

@login_required
def student_profile(request):
    if not hasattr(request.user, 'student_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
        
    student = request.user.student_profile
    if request.method == 'POST':
        form = StudentProfileForm(request.POST, request.FILES, instance=student)
        if form.is_valid():
            form.save()
            messages.success(request, 'Profile updated successfully.')
            return redirect('student_profile')
    else:
        form = StudentProfileForm(instance=student)
    
    context = {
        'student': student,
        'form': form
    }
    return render(request, 'accounts/student/profile.html', context)

@login_required
def student_profile_update(request):
    if not hasattr(request.user, 'student_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
    
    if request.method == 'POST':
        student = request.user.student_profile
        student.first_name = request.POST.get('first_name')
        student.last_name = request.POST.get('last_name')
        student.phone_number = request.POST.get('phone_number')
        student.date_of_birth = request.POST.get('date_of_birth')
        student.gender = request.POST.get('gender')
        student.address = request.POST.get('address')
        student.save()
        
        # Update the associated User model's first and last name
        user = request.user
        user.first_name = student.first_name
        user.last_name = student.last_name
        user.save()
        
        messages.success(request, 'Profile updated successfully.')
        return redirect('student_profile')
    
    return redirect('student_profile')

@login_required
def student_profile_picture_update(request):
    if not hasattr(request.user, 'student_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
    
    if request.method == 'POST' and request.FILES.get('photo'):
        student = request.user.student_profile
        student.photo = request.FILES['photo']
        student.save()
        messages.success(request, 'Profile picture updated successfully.')
    else:
        messages.error(request, 'Please select a photo to upload.')
    
    return redirect('student_profile')

@login_required
def student_courses(request):
    if not hasattr(request.user, 'student_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
        
    student = request.user.student_profile
    
    # Get enrolled courses
    enrolled_courses = student.enrolled_courses.all()
    
    # Add progress data to enrolled courses
    enrolled_courses_with_progress = []
    for course in enrolled_courses:
        course.progress = course.get_progress_for_student(student)
        enrolled_courses_with_progress.append(course)
    
    # Get available courses for enrollment
    # Use the new model method to filter courses the student can enroll in
    all_courses = Course.objects.all()
    available_courses = [course for course in all_courses if course.can_student_enroll(student)]
    
    context = {
        'current_courses': enrolled_courses_with_progress,
        'available_courses': available_courses,
        'student': student
    }
    return render(request, 'accounts/student/courses.html', context)

@login_required
def student_enroll_course(request, course_id):
    if not hasattr(request.user, 'student_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
        
    student = request.user.student_profile
    course = get_object_or_404(Course, id=course_id)
    
    # Check if student is already enrolled
    if student.enrolled_courses.filter(id=course_id).exists():
        messages.error(request, 'You are already enrolled in this course.')
        return redirect('student_courses')
    
    # Check if teacher is active
    if not course.teacher.is_active:
        messages.error(request, 'This course is not currently available for enrollment.')
        return redirect('student_courses')
    
    # Enroll the student
    student.enrolled_courses.add(course)
    course.students.add(student)
    
    messages.success(request, f'Successfully enrolled in {course.title}!')
    return redirect('student_courses')

@login_required
def student_timetable(request):
    if not hasattr(request.user, 'student_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
    
    from datetime import datetime, time as dtime
    student = request.user.student_profile
    schedules = Schedule.objects.filter(course__in=student.enrolled_courses.all()).select_related('course', 'course__teacher').order_by('day', 'start_time')

    # Always use admin defaults for periods
    start_hour = 8
    end_hour = 17
    slot_duration = 60

    # Build all possible time slots between start_hour and end_hour (exclusive)
    time_slots = []
    current_hour = start_hour
    current_minute = 0
    while current_hour < end_hour:
        start_time = dtime(current_hour, current_minute)
        end_minute = current_minute + slot_duration
        end_hour_calc = current_hour
        if end_minute >= 60:
            end_hour_calc += end_minute // 60
            end_minute = end_minute % 60
        end_time = dtime(end_hour_calc, end_minute)
        time_slots.append({
            'start': start_time,
            'end': end_time,
            'display': f"{start_time.strftime('%H:%M')} - {end_time.strftime('%H:%M')}"
        })
        current_minute += slot_duration
        if current_minute >= 60:
            current_hour += current_minute // 60
            current_minute = current_minute % 60
        else:
            current_hour += 0 if current_minute != 0 else 1

    days = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday']
    timetable_data = []
    for time_slot in time_slots:
        row_data = {
            'time_slot': time_slot,
            'cells': []
        }
        for day in days:
            schedule = schedules.filter(
                day=day,
                start_time__lte=time_slot['start'],
                end_time__gt=time_slot['start']
            ).first()
            row_data['cells'].append({
                'day': day,
                'schedule': schedule,
                'occupied': schedule is not None
            })
        timetable_data.append(row_data)

    # For summary count
    timetable = list(schedules)

    # Today's classes
    today_name = datetime.now().strftime('%A').lower()
    todays_classes = schedules.filter(day=today_name).order_by('start_time')
    todays_classes_list = []
    now = datetime.now().time()
    for c in todays_classes:
        is_current = c.start_time <= now < c.end_time
        is_completed = c.end_time <= now
        todays_classes_list.append({
            'course': c.course,
            'teacher': c.course.teacher,
            'room': c.room,
            'start_time': c.start_time.strftime('%H:%M'),
            'end_time': c.end_time.strftime('%H:%M'),
            'is_current': is_current,
            'is_completed': is_completed
        })

    # Upcoming events (next 5 for the student)
    upcoming_events = Event.objects.filter(
        start_date__gte=datetime.now()
    ).order_by('start_date')[:5]

    context = {
        'schedules': schedules,
        'start_hour': start_hour,
        'end_hour': end_hour,
        'slot_duration': slot_duration,
        'time_slots': time_slots,
        'timetable_data': timetable_data,
        'days': days,
        'upcoming_events': upcoming_events,
        'timetable': timetable,
        'todays_classes': todays_classes_list,
    }
    return render(request, 'accounts/student/timetable.html', context)

@login_required
def student_assignments(request):
    if not hasattr(request.user, 'student_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
        
    student = request.user.student_profile
    
    # Get all assignments for the student's enrolled courses
    assignments = Activity.objects.filter(
        course__enrolled_students=student,
        activity_type='assignment'
    ).select_related('course').order_by('-due_date')
    
    # Process each assignment to add submission and grade information
    assignment_data = []
    pending_count = 0
    submitted_count = 0
    graded_count = 0
    
    for assignment in assignments:
        # Check if student has submitted this assignment
        submission = AssignmentSubmission.objects.filter(
            student=student, 
            assignment=assignment
        ).first()
        
        # Check if assignment has been graded
        grade = Grade.objects.filter(
            student=student,
            subject=assignment.course,
            remarks__icontains=assignment.title
        ).first()
        
        # Determine status
        if grade:
            status = 'GRADED'
            status_color = 'success'
            graded_count += 1
        elif submission:
            status = 'SUBMITTED'
            status_color = 'info'
            submitted_count += 1
        else:
            status = 'PENDING'
            status_color = 'warning'
            pending_count += 1
        
        assignment_data.append({
            'id': assignment.id,
            'title': assignment.title,
            'description': assignment.description,
            'course': assignment.course,
            'due_date': assignment.due_date,
            'file': assignment.file,
            'status': status,
            'status_color': status_color,
            'submission': submission,
            'grade': grade,
            'submitted_at': submission.submitted_at if submission else None,
            'graded_at': grade.date if grade else None,
            'score': grade.score if grade else None,
            'grade_letter': grade.grade if grade else None
        })
    
    context = {
        'assignments': assignment_data,
        'pending_count': pending_count,
        'submitted_count': submitted_count,
        'graded_count': graded_count,
        'total_assignments': len(assignment_data)
    }
    return render(request, 'accounts/student/assignments.html', context)

@login_required
def student_assignment_detail(request, assignment_id):
    if not hasattr(request.user, 'student_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
        
    student = request.user.student_profile
    assignment = get_object_or_404(Activity, id=assignment_id, course__students=student)
    
    if request.method == 'POST':
        form = AssignmentSubmissionForm(request.POST, request.FILES)
        if form.is_valid():
            submission = form.save(commit=False)
            submission.student = student
            submission.assignment = assignment
            submission.save()
            messages.success(request, 'Assignment submitted successfully.')
            return redirect('student_assignments')
    else:
        form = AssignmentSubmissionForm()
    
    context = {
        'assignment': assignment,
        'form': form
    }
    return render(request, 'accounts/student/assignment_detail.html', context)

@login_required
def student_exams(request):
    if not hasattr(request.user, 'student_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
        
    student = request.user.student_profile
    today = datetime.now().date()
    
    # Get all exams for the student's enrolled courses
    all_exams = Examination.objects.filter(
        course__in=student.enrolled_courses.all()
    ).order_by('date')
    
    # Split exams into upcoming and completed
    upcoming_exams = all_exams.filter(date__gte=today)
    completed_exams = all_exams.filter(date__lt=today)
    
    # Calculate statistics
    total_completed = completed_exams.count()
    total_upcoming = upcoming_exams.count()
    
    # Get the next upcoming exam
    next_exam = upcoming_exams.first()
    
    exams_data = {
        'all': all_exams,
        'upcoming': upcoming_exams,
        'completed': completed_exams,
        'next': next_exam,
        'total_completed': total_completed,
        'total_upcoming': total_upcoming
    }

    # Attach schedule info to each exam
    for exam in all_exams:
        schedule = exam.examschedule_set.first()
        exam.start_time = schedule.start_time if schedule else None
        exam.venue = schedule.room if schedule else "Not assigned"

    context = {
        'exams': exams_data['all'],
        'today': today,
        'total_upcoming': total_upcoming,
        'total_completed': total_completed
    }
    return render(request, 'accounts/student/exams.html', context)

@login_required
def student_attendance(request):
    if not hasattr(request.user, 'student_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
        
    student = request.user.student_profile
    
    # Get current semester from request or default to current
    current_semester = request.GET.get('semester', '2024-1')
    available_semesters = ['2024-1', '2023-2', '2023-1']  # This should come from your database
    
    # Get all attendance records for the student
    attendance_records = Attendance.objects.filter(student=student)
    
    # Calculate overall attendance statistics
    total_classes = attendance_records.count()
    total_present = attendance_records.filter(status='present').count()
    total_absent = attendance_records.filter(status='absent').count()
    attendance_percentage = (total_present / total_classes * 100) if total_classes > 0 else 0
    
    # Calculate course-wise attendance
    course_attendance = []
    for course in student.enrolled_courses.all():
        course_records = attendance_records.filter(course=course)
        course_total = course_records.count()
        course_present = course_records.filter(status='present').count()
        course_absent = course_records.filter(status='absent').count()
        course_percentage = (course_present / course_total * 100) if course_total > 0 else 0
        
        course_attendance.append({
            'course': course,
            'total_classes': course_total,
            'present_count': course_present,
            'absent_count': course_absent,
            'attendance_percentage': course_percentage
        })
    
    context = {
        'student': student,
        'attendance_records': attendance_records,
        'total_classes': total_classes,
        'total_present': total_present,
        'total_absent': total_absent,
        'attendance_percentage': attendance_percentage,
        'course_attendance': course_attendance,
        'current_semester': current_semester,
        'available_semesters': available_semesters
    }
    return render(request, 'accounts/student/attendance.html', context)

@login_required
def student_fees(request):
    if not hasattr(request.user, 'student_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
        
    student = request.user.student_profile
    
    # Get all fee records for the student
    fees = Fee.objects.filter(student=student)
    invoices = Invoice.objects.filter(student=student).order_by('-created_at')
    payments = Payment.objects.filter(student=student).order_by('-payment_date')
    
    # Calculate fee summary
    total_fee = fees.aggregate(total=models.Sum('amount'))['total'] or 0
    paid_amount = payments.filter(status='completed').aggregate(total=models.Sum('amount'))['total'] or 0
    due_amount = total_fee - paid_amount
    
    # Get pending fees (unpaid fees)
    pending_fee = fees.filter(paid=False).order_by('due_date').first()
    
    # Get payment history with proper data structure including receipt download
    payment_history = []
    for payment in payments:
        payment_history.append({
            'id': payment.id,
            'transaction_id': payment.transaction_id,
            'date': payment.payment_date,
            'description': f"Fee Payment - {payment.payment_method}",
            'amount': payment.amount,
            'status': payment.status.upper(),
            'can_download_receipt': payment.status == 'completed',  # Only completed payments can download receipt
            'created_at': payment.created_at,
            'updated_at': payment.updated_at,
        })
    
    # Get successful payments for receipt download
    successful_payments = payments.filter(status='completed')
    
    context = {
        'total_fee': total_fee,
        'paid_amount': paid_amount,
        'due_amount': due_amount,
        'pending_fee': pending_fee,
        'payment_history': payment_history,
        'successful_payments': successful_payments,
        'invoices': invoices,
        'payments': payments
    }
    return render(request, 'accounts/student/fees.html', context)

@login_required
def student_make_payment(request):
    if not hasattr(request.user, 'student_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
    
    if request.method == 'POST':
        student = request.user.student_profile
        amount = request.POST.get('amount')
        payment_method = request.POST.get('payment_method')
        
        if not amount or not payment_method:
            messages.error(request, 'Please provide all required payment information.')
            return redirect('student_fees')
        
        try:
            # Create a new payment record
            payment = Payment.objects.create(
                student=student,
                amount=amount,
                payment_method=payment_method,
                status='PENDING',
                payment_date=datetime.now()
            )
            
            # Here you would typically integrate with a payment gateway
            # For now, we'll just mark it as successful
            payment.status = 'SUCCESS'
            payment.save()
            
            messages.success(request, 'Payment processed successfully.')
            return redirect('student_fees')
            
        except Exception as e:
            messages.error(request, f'An error occurred during payment processing: {str(e)}')
            return redirect('student_fees')
    
    return redirect('student_fees')

@login_required
def student_download_receipt(request):
    if not hasattr(request.user, 'student_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
    
    payment_id = request.GET.get('payment_id')
    if not payment_id:
        messages.error(request, 'No payment selected.')
        return redirect('student_fees')
    
    try:
        payment = Payment.objects.get(
            id=payment_id,
            student=request.user.student_profile,
            status='completed'
        )
        
        # Generate PDF receipt
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import inch
        from django.http import HttpResponse
        from datetime import datetime
        
        # Create HTTP response with PDF content type
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="receipt_{payment.transaction_id}.pdf"'
        
        # Create PDF
        p = canvas.Canvas(response, pagesize=A4)
        width, height = A4
        
        # Header
        p.setFont("Helvetica-Bold", 18)
        p.drawString(50, height - 80, "SARAT TECH - Payment Receipt")
        
        # Student and Payment Details
        p.setFont("Helvetica", 12)
        student = payment.student
        
        y_position = height - 150
        p.drawString(50, y_position, f"Student Name: {student.get_full_name()}")
        p.drawString(50, y_position - 25, f"Student ID: {student.student_id}")
        p.drawString(50, y_position - 50, f"Email: {student.email}")
        
        p.drawString(50, y_position - 90, f"Transaction ID: {payment.transaction_id}")
        p.drawString(50, y_position - 115, f"Payment Date: {payment.payment_date.strftime('%B %d, %Y')}")
        p.drawString(50, y_position - 140, f"Payment Method: {payment.payment_method}")
        p.drawString(50, y_position - 165, f"Amount: ${payment.amount}")
        p.drawString(50, y_position - 190, f"Status: {payment.status.upper()}")
        
        # Footer
        p.setFont("Helvetica", 10)
        p.drawString(50, 100, f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}")
        p.drawString(50, 80, "Thank you for your payment!")
        
        # Add a border
        p.rect(30, 50, width - 60, height - 100, stroke=1, fill=0)
        
        p.showPage()
        p.save()
        
        return response
        
    except Payment.DoesNotExist:
        messages.error(request, 'Payment not found or not eligible for receipt download.')
        return redirect('student_fees')
    except Exception as e:
        messages.error(request, f'Error generating receipt: {str(e)}')
        return redirect('student_fees')

@login_required
def student_notices(request):
    if not hasattr(request.user, 'student_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
        
    student = request.user.student_profile
    # Get all notices
    notices = list(Notice.objects.all())
    # Get all announcements for the student's enrolled courses
    announcements = list(Announcement.objects.filter(course__in=student.enrolled_courses.all(), is_active=True))
    # Combine and sort by created_at descending
    combined = notices + announcements
    combined_sorted = sorted(combined, key=lambda x: x.created_at, reverse=True)
    context = {
        'notices': combined_sorted
    }
    return render(request, 'accounts/student/notices.html', context)

@login_required
def teacher_dashboard(request):
    if not hasattr(request.user, 'teacher_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
        
    teacher = request.user.teacher_profile
    today = timezone.now().date()
    
    # Get teacher's courses
    my_courses = Course.objects.filter(teacher=teacher)
    for course in my_courses:
        course.active_student_count = course.students.filter(status='active').count()
    
    # Calculate total students as sum of students per course (not unique students)
    total_students = sum(course.students.filter(status='active').count() for course in my_courses)
    
    # Get unique students enrolled in teacher's courses
    my_students = Student.objects.filter(
        enrolled_courses__in=my_courses
    ).distinct()
    
    # Calculate detailed student statistics
    student_stats = {
        'total': my_students.count(),
        'active': my_students.filter(status='active').count(),
        'inactive': my_students.filter(status='inactive').count(),
        'male': my_students.filter(gender='male').count(),
        'female': my_students.filter(gender='female').count(),
    }
    
    # Calculate attendance statistics for students
    for student in my_students:
        # Calculate attendance percentage for this teacher's courses
        total_classes = Attendance.objects.filter(
            student=student,
            course__in=my_courses
        ).count()
        
        present_classes = Attendance.objects.filter(
            student=student,
            course__in=my_courses,
            status='present'
        ).count()
        
        student.attendance_percentage = round((present_classes / total_classes * 100), 1) if total_classes > 0 else 0
        
        # Calculate performance percentage based on grades
        grades = Grade.objects.filter(
            student=student,
            subject__in=my_courses
        )
        
        if grades.exists():
            total_score = sum(grade.score for grade in grades)
            total_possible = grades.count() * 100  # Assuming grades are out of 100
            student.performance = round((total_score / total_possible * 100), 1) if total_possible > 0 else 0
        else:
            student.performance = 0
        
        # Get recent attendance status
        recent_attendance = Attendance.objects.filter(
            student=student,
            course__in=my_courses,
            date__gte=today - timedelta(days=7)
        ).order_by('-date').first()
        
        student.recent_attendance = recent_attendance.status if recent_attendance else 'No data'
    
    # Today's schedule (Timetable)
    today_schedule = Schedule.objects.filter(
        course__teacher=teacher,
        day=today.strftime('%A').lower()
    ).order_by('start_time')
    
    # Attendance Overview (last 7 days)
    attendance_overview = []
    for i in range(6, -1, -1):
        day = today - timedelta(days=i)
        total = Attendance.objects.filter(course__teacher=teacher, date=day).count()
        present = Attendance.objects.filter(course__teacher=teacher, date=day, status='present').count()
        rate = round((present / total * 100), 1) if total > 0 else 0
        attendance_overview.append({'date': day, 'rate': rate})

    # Overall attendance rate (all time)
    all_attendance = Attendance.objects.filter(course__teacher=teacher)
    all_present = all_attendance.filter(status='present').count()
    attendance_rate = round((all_present / all_attendance.count() * 100), 1) if all_attendance.exists() else 0

    # Recent Activities (last 5 activities, assignments, or attendance records)
    recent_activities = list(Activity.objects.filter(course__teacher=teacher).order_by('-created_at')[:3])
    recent_activities += list(Attendance.objects.filter(course__teacher=teacher).order_by('-date')[:2])
    recent_activities = sorted(recent_activities, key=lambda x: getattr(x, 'created_at', getattr(x, 'date', None)), reverse=True)[:5]
    
    # Pending assignments
    pending_assignments = Activity.objects.filter(
        course__teacher=teacher,
        activity_type='assignment',
        due_date__gte=today
    ).order_by('due_date')[:5]
    
    # Upcoming exams
    upcoming_exams = Examination.objects.filter(
        course__teacher=teacher,
        date__gte=today
    ).order_by('date')[:5]
    
    # Top performing students
    top_students = sorted(my_students, key=lambda x: x.performance, reverse=True)[:5]
    
    # Students with low attendance (below 80%)
    low_attendance_students = [s for s in my_students if s.attendance_percentage < 80][:5]
    
    context = {
        'total_students': total_students,
        'active_courses': my_courses.count(),
        'classes_today': today_schedule.count(),
        'attendance_rate': attendance_rate,
        'today_schedule': today_schedule,
        'recent_activities': recent_activities,
        'my_courses': my_courses,
        'attendance_overview': attendance_overview,
        'my_students': my_students,
        'student_stats': student_stats,
        'top_students': top_students,
        'low_attendance_students': low_attendance_students,
        'pending_assignments': pending_assignments,
        'upcoming_exams': upcoming_exams,
    }
    return render(request, 'accounts/teacher/dashboard.html', context)

@login_required
def teacher_profile(request):
    if not hasattr(request.user, 'teacher_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
        
    teacher = request.user.teacher_profile
    if request.method == 'POST':
        form = TeacherProfileForm(request.POST, request.FILES, instance=teacher)
        if form.is_valid():
            form.save()
            messages.success(request, 'Profile updated successfully.')
            return redirect('teacher_profile')
    else:
        form = TeacherProfileForm(instance=teacher)
    
    context = {
        'teacher': teacher,
        'form': form
    }
    return render(request, 'accounts/teacher/profile/overview.html', context)

@login_required
def teacher_profile_update(request):
    if not hasattr(request.user, 'teacher_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
    
    if request.method == 'POST':
        teacher = request.user.teacher_profile
        teacher.first_name = request.POST.get('first_name')
        teacher.last_name = request.POST.get('last_name')
        teacher.phone_number = request.POST.get('phone_number')
        teacher.date_of_birth = request.POST.get('date_of_birth')
        teacher.gender = request.POST.get('gender')
        teacher.address = request.POST.get('address')
        teacher.department = request.POST.get('department')
        teacher.qualification = request.POST.get('qualification')
        teacher.save()
        
        # Update the associated User model's first and last name
        user = request.user
        user.first_name = teacher.first_name
        user.last_name = teacher.last_name
        user.save()
        
        messages.success(request, 'Profile updated successfully.')
        return redirect('teacher_profile')
    
    return redirect('teacher_profile')

@login_required
def teacher_profile_picture_update(request):
    if not hasattr(request.user, 'teacher_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
    
    if request.method == 'POST' and request.FILES.get('photo'):
        teacher = request.user.teacher_profile
        teacher.photo = request.FILES['photo']
        teacher.save()
        messages.success(request, 'Profile picture updated successfully.')
    else:
        messages.error(request, 'Please select a photo to upload.')
    
    return redirect('teacher_profile')

@login_required
def teacher_courses(request):
    if not hasattr(request.user, 'teacher_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
        
    teacher = request.user.teacher_profile
    courses = Course.objects.filter(teacher=teacher)
    return render(request, 'accounts/teacher/courses/list.html', {'courses': courses})

@login_required
def teacher_students(request):
    if not hasattr(request.user, 'teacher_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
        
    teacher = request.user.teacher_profile
    
    # Get all courses taught by the teacher
    courses = Course.objects.filter(teacher=teacher)
    
    # Get all classes that have students enrolled in the teacher's courses
    classes = Class.objects.filter(
        students__enrolled_courses__in=courses
    ).distinct()
    
    # Get all students enrolled in the teacher's courses
    students = Student.objects.filter(
        enrolled_courses__in=courses
    ).distinct()
    
    # Calculate attendance and performance for each student
    for student in students:
        # Calculate attendance percentage
        total_classes = Attendance.objects.filter(
            student=student,
            course__in=courses
        ).count()
        
        present_classes = Attendance.objects.filter(
            student=student,
            course__in=courses,
            status='present'
        ).count()
        
        student.attendance_percentage = (present_classes / total_classes * 100) if total_classes > 0 else 0
        
        # Calculate performance percentage based on grades
        grades = Grade.objects.filter(
            student=student,
            subject__in=courses
        )
        
        if grades.exists():
            total_score = sum(grade.score for grade in grades)
            total_possible = grades.count() * 100  # Assuming grades are out of 100
            student.performance = (total_score / total_possible * 100) if total_possible > 0 else 0
        else:
            student.performance = 0
    
    context = {
        'courses': courses,
        'classes': classes,
        'students': students
    }
    return render(request, 'accounts/teacher/students/list.html', context)

@login_required
def teacher_assignments(request):
    if not hasattr(request.user, 'teacher_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
        
    teacher = request.user.teacher_profile
    
    # Get assignments with submission counts and student counts
    assignments = Activity.objects.filter(
        course__teacher=teacher,
        activity_type='assignment'
    ).select_related('course').prefetch_related(
        'submissions',
        'course__enrolled_students'
    ).order_by('-created_at')
    
    # Add submission and student counts to each assignment
    assignments_with_counts = []
    for assignment in assignments:
        submission_count = assignment.submissions.count()
        total_students = assignment.course.enrolled_students.count()
        
        assignment.submission_count = submission_count
        assignment.total_students = total_students
        assignment.submission_percentage = int((submission_count / total_students * 100)) if total_students > 0 else 0
        
        assignments_with_counts.append(assignment)
    
    courses = Course.objects.filter(teacher=teacher)
    
    context = {
        'assignments': assignments_with_counts,
        'courses': courses,
        'today': timezone.now().date()
    }
    return render(request, 'accounts/teacher/assignments/list.html', context)

@login_required
def teacher_create_assignment(request):
    if not hasattr(request.user, 'teacher_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
        
    teacher = request.user.teacher_profile
    
    if request.method == 'POST':
        course_id = request.POST.get('course')
        title = request.POST.get('title')
        description = request.POST.get('description')
        due_date = request.POST.get('due_date')
        max_score = request.POST.get('max_score')
        
        try:
            course = Course.objects.get(id=course_id, teacher=teacher)
            assignment = Activity.objects.create(
                course=course,
                title=title,
                description=description,
                due_date=due_date,
                max_score=max_score,
                activity_type='assignment'
            )
            messages.success(request, 'Assignment created successfully.')
            return redirect('teacher_assignments')
        except Course.DoesNotExist:
            messages.error(request, 'Invalid course selected.')
            return redirect('teacher_assignments')
    
    # This view should only handle POST requests for the modal form
    return redirect('teacher_assignments')

@login_required
def teacher_assignment_detail(request, assignment_id):
    if not hasattr(request.user, 'teacher_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
        
    teacher = request.user.teacher_profile
    assignment = get_object_or_404(Activity, id=assignment_id, course__teacher=teacher)
    
    # Get students enrolled in the course for this assignment
    enrolled_students = assignment.course.enrolled_students.all()
    
    context = {
        'assignment': assignment,
        'enrolled_students': enrolled_students
    }
    return render(request, 'accounts/teacher/assignment_detail.html', context)

@login_required
def teacher_assignment_submissions(request, assignment_id):
    if not hasattr(request.user, 'teacher_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
        
    teacher = request.user.teacher_profile
    assignment = get_object_or_404(Activity, id=assignment_id, course__teacher=teacher)
    
    # Get all submissions for this assignment
    submissions = AssignmentSubmission.objects.filter(assignment=assignment).select_related('student').order_by('-submitted_at')
    
    # Get all students enrolled in the course
    enrolled_students = assignment.course.enrolled_students.all()
    
    # Calculate submission statistics
    total_students = enrolled_students.count()
    submitted_count = submissions.count()
    pending_count = total_students - submitted_count
    submission_percentage = int((submitted_count / total_students * 100)) if total_students > 0 else 0
    
    # Create submission data with grades
    submission_data = []
    for submission in submissions:
        # Get grade for this assignment and student
        grade = Grade.objects.filter(
            student=submission.student,
            subject=assignment.course,
            remarks__icontains=assignment.title
        ).first()
        
        submission_data.append({
            'submission': submission,
            'grade': grade,
            'status': 'Submitted',
            'status_color': 'success'
        })
    
    # Add students who haven't submitted
    submitted_student_ids = [s.student.id for s in submissions]
    for student in enrolled_students:
        if student.id not in submitted_student_ids:
            submission_data.append({
                'submission': None,
                'student': student,
                'grade': None,
                'status': 'Pending',
                'status_color': 'warning'
            })
    
    context = {
        'assignment': assignment,
        'submissions': submission_data,
        'total_students': total_students,
        'submitted_count': submitted_count,
        'pending_count': pending_count,
        'submission_percentage': submission_percentage,
    }
    return render(request, 'accounts/teacher/assignments/submissions.html', context)

@login_required
def teacher_view_submission(request, assignment_id, submission_id):
    if not hasattr(request.user, 'teacher_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
        
    teacher = request.user.teacher_profile
    assignment = get_object_or_404(Activity, id=assignment_id, course__teacher=teacher)
    submission = get_object_or_404(AssignmentSubmission, id=submission_id, assignment=assignment)
    
    # Get existing grade for this submission
    grade = Grade.objects.filter(
        student=submission.student,
        subject=assignment.course,
        remarks__icontains=assignment.title
    ).first()
    
    context = {
        'assignment': assignment,
        'submission': submission,
        'grade': grade,
    }
    return render(request, 'accounts/teacher/assignments/view_submission.html', context)

@login_required
def teacher_grade_submission(request, assignment_id, submission_id):
    if not hasattr(request.user, 'teacher_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
        
    teacher = request.user.teacher_profile
    assignment = get_object_or_404(Activity, id=assignment_id, course__teacher=teacher)
    submission = get_object_or_404(AssignmentSubmission, id=submission_id, assignment=assignment)
    
    if request.method == 'POST':
        try:
            score = float(request.POST.get('score', 0))
            grade_letter = request.POST.get('grade_letter', '')
            remarks = request.POST.get('remarks', f'Grade for {assignment.title}')
            
            # Create or update grade
            grade, created = Grade.objects.get_or_create(
                student=submission.student,
                subject=assignment.course,
                remarks=f'Assignment: {assignment.title}',
                defaults={
                    'score': score,
                    'grade': grade_letter,
                    'date': timezone.now().date(),
                }
            )
            
            if not created:
                grade.score = score
                grade.grade = grade_letter
                grade.remarks = remarks
                grade.date = timezone.now().date()
                grade.save()
            
            messages.success(request, f'Grade saved successfully for {submission.student.get_full_name()}.')
            return redirect('teacher_assignment_submissions', assignment_id=assignment_id)
            
        except Exception as e:
            messages.error(request, f'Error saving grade: {str(e)}')
    
    return redirect('teacher_view_submission', assignment_id=assignment_id, submission_id=submission_id)

@login_required
def teacher_edit_assignment(request, assignment_id):
    if not hasattr(request.user, 'teacher_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
        
    teacher = request.user.teacher_profile
    assignment = get_object_or_404(Activity, id=assignment_id, course__teacher=teacher)
    
    if request.method == 'POST':
        try:
            course_id = request.POST.get('course')
            title = request.POST.get('title')
            description = request.POST.get('description')
            due_date = request.POST.get('due_date')
            max_score = request.POST.get('max_score')
            
            # Update assignment
            if course_id:
                course = Course.objects.get(id=course_id, teacher=teacher)
                assignment.course = course
            
            assignment.title = title
            assignment.description = description
            assignment.due_date = due_date
            assignment.max_score = max_score
            assignment.save()
            
            messages.success(request, 'Assignment updated successfully.')
            return redirect('teacher_assignments')
            
        except Course.DoesNotExist:
            messages.error(request, 'Invalid course selected.')
        except Exception as e:
            messages.error(request, f'Error updating assignment: {str(e)}')
    
    # Get teacher's courses for the form
    courses = Course.objects.filter(teacher=teacher)
    
    context = {
        'assignment': assignment,
        'courses': courses,
        'today': timezone.now().date()
    }
    return render(request, 'accounts/teacher/assignments/edit.html', context)

@login_required
def teacher_delete_assignment(request, assignment_id):
    if not hasattr(request.user, 'teacher_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
        
    teacher = request.user.teacher_profile
    assignment = get_object_or_404(Activity, id=assignment_id, course__teacher=teacher)
    
    if request.method == 'POST':
        try:
            assignment_title = assignment.title
            # Delete associated submissions first
            AssignmentSubmission.objects.filter(assignment=assignment).delete()
            # Delete associated grades
            Grade.objects.filter(
                subject=assignment.course,
                remarks__icontains=assignment.title
            ).delete()
            # Delete the assignment
            assignment.delete()
            
            messages.success(request, f'Assignment "{assignment_title}" has been deleted successfully.')
        except Exception as e:
            messages.error(request, f'Error deleting assignment: {str(e)}')
    
    return redirect('teacher_assignments')

@login_required
def teacher_attendance_index(request):
    if not hasattr(request.user, 'teacher_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
        
    teacher = request.user.teacher_profile
    today = datetime.now().date()
    
    # Get teacher's courses
    courses = Course.objects.filter(teacher=teacher)
    
    # Get all classes for the teacher's courses
    classes = Class.objects.filter(courses__teacher=teacher).distinct()
    
    # Calculate today's attendance statistics
    today_attendance = Attendance.objects.filter(
        course__teacher=teacher,
        date=today
    )
    today_attendance_count = today_attendance.count()
    present_today_count = today_attendance.filter(status='present').count()
    absent_today_count = today_attendance.filter(status='absent').count()
    
    # Get total classes
    total_classes = Class.objects.filter(courses__teacher=teacher).distinct().count()
    
    # Get recent attendance records (last 5)
    recent_attendance = []
    for course in courses:
        attendance = Attendance.objects.filter(
            course=course,
            date__lte=today
        ).order_by('-date')[:5]
        
        for record in attendance:
            total_students = course.students.count()
            present_count = Attendance.objects.filter(
                course=course,
                date=record.date,
                status='present'
            ).count()
            absent_count = total_students - present_count
            
            recent_attendance.append({
                'date': record.date,
                'course': course,
                'total_students': total_students,
                'present_count': present_count,
                'absent_count': absent_count
            })
    
    # Sort recent attendance by date (newest first)
    recent_attendance.sort(key=lambda x: x['date'], reverse=True)
    recent_attendance = recent_attendance[:5]  # Limit to 5 most recent records
    
    context = {
        'today_attendance_count': today_attendance_count,
        'present_today_count': present_today_count,
        'absent_today_count': absent_today_count,
        'total_classes': total_classes,
        'recent_attendance': recent_attendance,
        'courses': courses,
        'classes': classes
    }
    return render(request, 'accounts/teacher/attendance/index.html', context)

@login_required
def teacher_attendance(request):
    if not hasattr(request.user, 'teacher_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
    
    teacher = request.user.teacher_profile
    courses = Course.objects.filter(teacher=teacher)
    today = datetime.now().date()
    classes = Class.objects.filter(courses__teacher=teacher).distinct()
    
    # Filtering
    course_id = request.GET.get('course')
    class_id = request.GET.get('class')
    date_str = request.GET.get('date')
    
    thirty_days_ago = today - timedelta(days=30)
    attendance_qs = Attendance.objects.filter(
        course__teacher=teacher,
        date__gte=thirty_days_ago,
        date__lte=today
    ).order_by('-date')
    if course_id:
        attendance_qs = attendance_qs.filter(course_id=course_id)
    if class_id:
        attendance_qs = attendance_qs.filter(course__class_section_id=class_id)
    if date_str:
        try:
            filter_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            attendance_qs = attendance_qs.filter(date=filter_date)
        except Exception:
            pass
    # Group by (date, course)
    grouped = {}
    for record in attendance_qs:
        key = (record.date, record.course.id)
        if key not in grouped:
            grouped[key] = {
                'date': record.date,
                'course': record.course,
                'total_students': record.course.students.count(),
                'present_count': 0,
                'absent_count': 0
            }
        if record.status == 'present':
            grouped[key]['present_count'] += 1
        else:
            grouped[key]['absent_count'] += 1
    attendance_records = list(grouped.values())
    attendance_records.sort(key=lambda x: x['date'], reverse=True)
    # Pagination
    paginator = Paginator(attendance_records, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    # Attendance summary for each course
    attendance_summary = []
    for course in courses:
        total_students = course.students.count()
        present_today = Attendance.objects.filter(
            course=course,
            date=today,
            status='present'
        ).count()
        attendance_summary.append({
            'course': course,
            'total_students': total_students,
            'present_today': present_today,
            'attendance_percentage': (present_today/total_students*100) if total_students > 0 else 0
        })
    context = {
        'courses': courses,
        'classes': classes,
        'attendance_records': page_obj,
        'attendance_summary': attendance_summary,
        'today': today
    }
    return render(request, 'accounts/teacher/attendance/history.html', context)

@login_required
def teacher_attendance_select(request):
    if not hasattr(request.user, 'teacher_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
        
    teacher = request.user.teacher_profile
    courses = Course.objects.filter(teacher=teacher)
    
    context = {
        'courses': courses
    }
    return render(request, 'accounts/teacher/attendance/select.html', context)

@login_required
def teacher_take_attendance(request, course_id):
    if not hasattr(request.user, 'teacher_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
        
    teacher = request.user.teacher_profile
    course = get_object_or_404(Course, id=course_id, teacher=teacher)
    today = datetime.now().date()
    
    if request.method == 'POST':
        missing_status = False
        for student_id in request.POST.getlist('student_ids'):
            status = request.POST.get(f'status_{student_id}')
            if not status:
                missing_status = True
                break
        if missing_status:
            messages.error(request, 'Please select attendance status for all students before submitting.')
            return redirect(request.path)
        for student_id in request.POST.getlist('student_ids'):
            status = request.POST.get(f'status_{student_id}')
            student = Student.objects.get(id=student_id)
            # Update or create attendance record
            Attendance.objects.update_or_create(
                student=student,
                course=course,
                date=today,
                defaults={'status': status}
            )
        messages.success(request, 'Attendance recorded successfully.')
        return redirect('teacher_attendance')
    
    students = course.students.all()
    attendance_records = {
        record.student_id: record.status 
        for record in Attendance.objects.filter(course=course, date=today)
    }
    
    context = {
        'course': course,
        'students': students,
        'attendance_records': attendance_records,
        'today': today
    }
    return render(request, 'accounts/teacher/attendance/take.html', context)

@login_required
def teacher_exams(request):
    if not hasattr(request.user, 'teacher_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
        
    teacher = request.user.teacher_profile
    exams = Examination.objects.filter(course__teacher=teacher).order_by('-date')
    courses = Course.objects.filter(teacher=teacher)
    
    context = {
        'exams': exams,
        'courses': courses
    }
    return render(request, 'accounts/teacher/exams/list.html', context)

@login_required
def teacher_create_exam(request):
    if not hasattr(request.user, 'teacher_profile'):
        messages.error(request, 'You do not have access to this feature.')
        return redirect('login')
        
    teacher = request.user.teacher_profile
    
    if request.method == 'POST':
        course_id = request.POST.get('course')
        title = request.POST.get('title')
        exam_type = request.POST.get('exam_type')
        date_str = request.POST.get('date')
        duration = request.POST.get('duration')
        total_marks = request.POST.get('total_marks')
        description = request.POST.get('description')

        from datetime import datetime, timedelta
        try:
            course = Course.objects.get(id=course_id, teacher=teacher)
            # Parse date and duration
            exam_date = datetime.strptime(date_str, '%Y-%m-%dT%H:%M') if date_str else None
            exam_duration = timedelta(minutes=int(duration)) if duration else None
            exam = Examination.objects.create(
                course=course,
                title=title,
                exam_type=exam_type,
                date=exam_date,
                duration=exam_duration,
                total_marks=total_marks,
                instructions=description
            )
            messages.success(request, 'Exam created successfully.')
            return redirect('teacher_exam_detail', exam_id=exam.id)
        except Course.DoesNotExist:
            messages.error(request, 'Invalid course selected.')
        except Exception as e:
            messages.error(request, f'Error creating exam: {str(e)}')
    
    courses = Course.objects.filter(teacher=teacher)
    context = {
        'courses': courses
    }
    return render(request, 'accounts/teacher/exams/list.html', context)

@login_required
def teacher_exam_detail(request, exam_id):
    if not hasattr(request.user, 'teacher_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
        
    teacher = request.user.teacher_profile
    exam = get_object_or_404(Examination, id=exam_id, course__teacher=teacher)
    students = exam.course.students.all()
    
    if request.method == 'POST':
        for student in students:
            marks = request.POST.get(f'marks_{student.id}')
            if marks:
                Grade.objects.update_or_create(
                    student=student,
                    subject=exam.course,
                    examination=exam,
                    defaults={'score': marks, 'date': exam.date}
                )
        messages.success(request, 'Marks updated successfully.')
        return redirect('teacher_exam_detail', exam_id=exam.id)
    
    results = {
        grade.student_id: grade.score 
        for grade in Grade.objects.filter(examination=exam)
    }
    
    context = {
        'exam': exam,
        'students': students,
        'results': results
    }
    return render(request, 'accounts/teacher/exam_detail.html', context)

@login_required
def teacher_grades(request):
    if not hasattr(request.user, 'teacher_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
    
    teacher = request.user.teacher_profile
    courses = Course.objects.filter(teacher=teacher).select_related('class_section')
    
    context = {
        'courses': courses,
    }
    return render(request, 'accounts/teacher/exams/grades.html', context)

@login_required
def teacher_course_grades(request, course_id):
    if not hasattr(request.user, 'teacher_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
    try:
        teacher = request.user.teacher_profile
        print(f"DEBUG: teacher_course_grades called for course_id={course_id}, teacher={teacher}")
        
        # Get course with related data
        course = get_object_or_404(Course, id=course_id, teacher=teacher)
        
        # Get enrolled students (using the correct relationship)
        students = course.enrolled_students.filter(status='active').order_by('student_id')
        
        # Get all exams and assignments for this course
        exams = Examination.objects.filter(course=course).order_by('date')
        assignments = Activity.objects.filter(course=course, activity_type='assignment').order_by('due_date')
        
        print(f"DEBUG: Course: {course}, Students: {students.count()}, Exams: {exams.count()}, Assignments: {assignments.count()}")
        
        student_grades = []
        for student in students:
            # Get all actual Grade records for this student in this course
            student_grades_for_course = Grade.objects.filter(
                student=student, 
                subject=course
            ).select_related('subject').order_by('-date')
            
            # Calculate totals from actual grades
            if student_grades_for_course.exists():
                total_score = sum(float(grade.score) for grade in student_grades_for_course)
                total_possible = student_grades_for_course.count() * 100
                grade_percentage = (total_score / total_possible * 100) if total_possible > 0 else 0
                avg_score = total_score / student_grades_for_course.count()
            else:
                total_score = 0
                grade_percentage = 0
                avg_score = 0
            
            # Get assignment data (for display purposes)
            assignment_scores = []
            for assignment in assignments:
                submission = AssignmentSubmission.objects.filter(
                    student=student,
                    assignment=assignment
                ).first()
                
                # Check if there's a grade specifically for this assignment
                assignment_grade = student_grades_for_course.filter(
                    remarks__icontains=assignment.title
                ).first()
                
                score = assignment_grade.score if assignment_grade else (50 if submission else 0)
                max_score = getattr(assignment, 'max_score', 100)
                assignment_scores.append({
                    'activity': assignment, 
                    'score': score,
                    'max_score': max_score,
                    'submitted': bool(submission)
                })
            
            # Get exam data (for display purposes) 
            exam_scores = []
            for exam in exams:
                exam_grade = student_grades_for_course.filter(
                    examination=exam
                ).first()
                marks = exam_grade.score if exam_grade else 0
                exam_scores.append({
                    'examination': exam, 
                    'marks': marks,
                    'total_marks': exam.total_marks
                })
            
            # Calculate attendance percentage for this course
            total_attendance = Attendance.objects.filter(student=student, course=course).count()
            present_attendance = Attendance.objects.filter(
                student=student, 
                course=course, 
                status='present'
            ).count()
            attendance_percentage = (present_attendance / total_attendance * 100) if total_attendance > 0 else 0
            
            student_grades.append({
                'student': student,
                'grades': student_grades_for_course,
                'total_score': round(total_score, 2),
                'grade_percentage': round(grade_percentage, 2),
                'avg_score': round(avg_score, 2),
                'total_grade_records': student_grades_for_course.count(),
                'assignment_scores': assignment_scores,
                'exam_scores': exam_scores,
                'attendance_percentage': round(attendance_percentage, 2),
                'total_assignments': len(assignment_scores),
                'completed_assignments': sum(1 for a in assignment_scores if a['submitted']),
                'total_exams': len(exam_scores),
                'graded_exams': sum(1 for e in exam_scores if e['marks'] > 0)
            })
        
        print(f"DEBUG: Processed {len(student_grades)} students with grades")
        
        # Course statistics
        course_stats = {
            'total_students': students.count(),
            'total_assignments': assignments.count(),
            'total_exams': exams.count(),
            'avg_grade': round(sum(sg['grade_percentage'] for sg in student_grades) / len(student_grades), 2) if student_grades else 0,
            'highest_grade': max((sg['grade_percentage'] for sg in student_grades), default=0),
            'lowest_grade': min((sg['grade_percentage'] for sg in student_grades), default=0)
        }
        
        context = {
            'course': course,
            'student_grades': student_grades,
            'course_stats': course_stats,
            'assignments': assignments,
            'exams': exams
        }
        return render(request, 'accounts/teacher/course_grades.html', context)
        
    except Exception as e:
        print(f"ERROR in teacher_course_grades: {e}")
        import traceback
        traceback.print_exc()
        messages.error(request, f"An error occurred loading course grades: {str(e)}")
        context = {
            'course': get_object_or_404(Course, id=course_id) if course_id else None,
            'student_grades': [],
            'course_stats': {},
            'assignments': [],
            'exams': []
        }
        return render(request, 'accounts/teacher/course_grades.html', context)

@login_required
def teacher_timetable(request):
    if not hasattr(request.user, 'teacher_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
        
    teacher = request.user.teacher_profile
    
    # Get URL parameters for time settings
    start_hour = int(request.GET.get('start_hour', 8))
    end_hour = int(request.GET.get('end_hour', 17))
    slot_duration = int(request.GET.get('duration', 60))
    
    # Validate parameters
    if start_hour >= end_hour:
        start_hour, end_hour = 8, 17
    
    # Generate time slots
    time_slots = []
    current_time = start_hour * 60  # Convert to minutes
    end_minutes = end_hour * 60
    
    while current_time < end_minutes:
        start_minutes = current_time
        end_minutes_slot = min(current_time + slot_duration, end_minutes)
        
        start_time = f"{start_minutes // 60:02d}:{start_minutes % 60:02d}"
        end_time = f"{end_minutes_slot // 60:02d}:{end_minutes_slot % 60:02d}"
        
        time_slots.append({
            'start': start_time,
            'end': end_time,
            'display': f"{start_time} - {end_time}"
        })
        
        current_time += slot_duration
    
    # Define days
    days = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday']
    
    # Get all schedules for the teacher's courses
    schedules = Schedule.objects.filter(course__teacher=teacher).order_by('day', 'start_time')
    
    # Get teacher's courses for the form
    courses = Course.objects.filter(teacher=teacher)
    
    # Get all classes for filtering
    classes = Class.objects.all()
    
    # Create timetable data structure
    timetable_data = []
    for time_slot in time_slots:
        row = {
            'time_slot': time_slot,
            'cells': []
        }
        
        for day in days:
            cell_schedules = schedules.filter(
                day=day,
                start_time__lte=time_slot['start'],
                end_time__gt=time_slot['start']
            )
            
            cell = {
                'day': day,
                'occupied': cell_schedules.exists(),
                'schedule': cell_schedules.first() if cell_schedules.exists() else None
            }
            row['cells'].append(cell)
        
        timetable_data.append(row)
    
    # Handle form submission for adding new schedule
    if request.method == 'POST':
        form = ScheduleForm(request.POST)
        if form.is_valid():
            try:
                # Check if the course belongs to the teacher
                course = form.cleaned_data['course']
                if course.teacher != teacher:
                    return JsonResponse({
                        'success': False,
                        'errors': {'general': ['You can only schedule your own courses.']}
                    })
                
                # Check for conflicts
                new_schedule = form.save(commit=False)
                conflicts = Schedule.objects.filter(
                    day=new_schedule.day,
                    room=new_schedule.room,
                    start_time__lt=new_schedule.end_time,
                    end_time__gt=new_schedule.start_time
                )
                
                if conflicts.exists():
                    return JsonResponse({
                        'success': False,
                        'errors': {'general': ['This time slot conflicts with an existing schedule.']}
                    })
                
                # Save the schedule
                new_schedule.save()
                
                # Return success response with schedule data
                return JsonResponse({
                    'success': True,
                    'message': 'Class scheduled successfully!',
                    'schedule': {
                        'id': new_schedule.id,
                        'course_title': new_schedule.course.title,
                        'teacher_name': new_schedule.course.teacher.get_full_name(),
                        'room': new_schedule.room,
                        'day': new_schedule.day,
                        'start_time': new_schedule.start_time.strftime('%H:%M'),
                        'end_time': new_schedule.end_time.strftime('%H:%M')
                    }
                })
                
            except Exception as e:
                print(f"Error saving schedule: {e}")
                return JsonResponse({
                    'success': False,
                    'errors': {'general': ['An error occurred while saving the schedule.']}
                })
        else:
            return JsonResponse({
                'success': False,
                'errors': form.errors
            })
    else:
        form = ScheduleForm()
        # Filter courses to only show teacher's courses
        form.fields['course'].queryset = courses
    
    context = {
        'timetable_data': timetable_data,
        'schedules': schedules,
        'courses': courses,
        'classes': classes,
        'days': days,
        'time_slots': time_slots,
        'start_hour': start_hour,
        'end_hour': end_hour,
        'slot_duration': slot_duration,
        'form': form
    }
    
    return render(request, 'accounts/teacher/timetable.html', context)

@login_required
def teacher_course_students(request, course_id):
    if not hasattr(request.user, 'teacher_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
        
    teacher = request.user.teacher_profile
    course = get_object_or_404(Course, id=course_id, teacher=teacher)
    students = course.students.all()
    
    context = {
        'course': course,
        'students': students
    }
    return render(request, 'accounts/teacher/course_students.html', context)

@login_required
def teacher_course_assignments(request, course_id):
    if not hasattr(request.user, 'teacher_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
        
    teacher = request.user.teacher_profile
    course = get_object_or_404(Course, id=course_id, teacher=teacher)
    assignments = Activity.objects.filter(
        course=course,
        activity_type='assignment'
    ).order_by('-created_at')
    
    context = {
        'course': course,
        'assignments': assignments
    }
    return render(request, 'accounts/teacher/course_assignments.html', context)

@login_required
def teacher_create_notice(request):
    if not hasattr(request.user, 'teacher_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
        
    teacher = request.user.teacher_profile
    
    if request.method == 'POST':
        title = request.POST.get('title')
        content = request.POST.get('content')
        notice_type = request.POST.get('notice_type', 'academic')
        target_audience = request.POST.get('target_audience', 'students')
        attachment = request.FILES.get('attachment')
        
        
        notice = Notice.objects.create(
            title=title,
            content=content,
            notice_type=notice_type,
            target_audience=target_audience,
            attachment=attachment,
            created_by=request.user,
            is_active=True
        )
        
        messages.success(request, 'Notice created successfully.')
        return redirect('teacher_dashboard')
    
    context = {
        'notice_types': Notice.NOTICE_TYPES,
        'target_audiences': [
            ('students', 'Students'),
            ('parents', 'Parents'),
            ('all', 'All')
        ]
    }
    return render(request, 'accounts/teacher/create_notice.html', context)

@login_required
def teacher_messages(request):
    if not hasattr(request.user, 'teacher_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
        
    teacher = request.user.teacher_profile
    messages_list = Message.objects.filter(
        Q(sender=request.user) | Q(recipient=request.user)
    ).order_by('-created_at')
    
    # Get selected message if any
    message_id = request.GET.get('message_id')
    selected_message = None
    if message_id:
        try:
            selected_message = Message.objects.get(id=message_id)
            # Mark message as read if recipient is viewing it
            if selected_message.recipient == request.user and not selected_message.is_read:
                selected_message.is_read = True
                selected_message.save()
        except Message.DoesNotExist:
            pass
    
    context = {
        'messages': messages_list,
        'selected_message': selected_message
    }
    return render(request, 'accounts/teacher/communication/messages.html', context)

@login_required
def teacher_announcements(request):
    if not hasattr(request.user, 'teacher_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
        
    teacher = request.user.teacher_profile
    courses = Course.objects.filter(teacher=teacher)
    announcements = Announcement.objects.filter(course__in=courses).order_by('-created_at')
    
    context = {
        'announcements': announcements,
        'courses': courses
    }
    return render(request, 'accounts/teacher/communication/announcements.html', context)

@login_required
def teacher_create_announcement(request):
    if not hasattr(request.user, 'teacher_profile') or request.method != 'POST':
        messages.error(request, 'Invalid request.')
        return redirect('teacher_announcements')
    
    teacher = request.user.teacher_profile
    
    try:
        # Get form data
        title = request.POST.get('title')
        content = request.POST.get('content')
        category = request.POST.get('category')
        course_id = request.POST.get('course')
        expiry_date = request.POST.get('expiry_date')
        
        # Get the course and verify it belongs to the teacher
        course = get_object_or_404(Course, id=course_id, teacher=teacher)
        
        # Create announcement
        announcement = Announcement.objects.create(
            course=course,
            category=category,
            title=title,
            content=content,
            created_by=request.user,
            expiry_date=expiry_date if expiry_date else None
        )
        
        # Handle attachment if provided
        if 'attachment' in request.FILES:
            attachment_file = request.FILES['attachment']
            
            # Create attachment
            AnnouncementAttachment.objects.create(
                announcement=announcement,
                file=attachment_file,
                filename=attachment_file.name,
                file_size=attachment_file.size
            )
        
        messages.success(request, 'Announcement created successfully.')
        
    except Exception as e:
        messages.error(request, f'Error creating announcement: {str(e)}')
    
    return redirect('teacher_announcements')

@login_required
def teacher_edit_announcement(request, announcement_id):
    if not hasattr(request.user, 'teacher_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
    
    teacher = request.user.teacher_profile
    announcement = get_object_or_404(Announcement, id=announcement_id, created_by=request.user)
    
    if request.method == 'POST':
        try:
            # Update announcement fields
            announcement.title = request.POST.get('title')
            announcement.content = request.POST.get('content')
            announcement.category = request.POST.get('category')
            
            # Update course if provided and verify it belongs to the teacher
            course_id = request.POST.get('course')
            if course_id:
                course = get_object_or_404(Course, id=course_id, teacher=teacher)
                announcement.course = course
            
            expiry_date = request.POST.get('expiry_date')
            announcement.expiry_date = expiry_date if expiry_date else None
            
            announcement.save()
            
            # Handle new attachment if provided
            if 'attachment' in request.FILES:
                # Remove old attachments
                announcement.announcementattachment_set.all().delete()
                
                attachment_file = request.FILES['attachment']
                AnnouncementAttachment.objects.create(
                    announcement=announcement,
                    file=attachment_file,
                    filename=attachment_file.name,
                    file_size=attachment_file.size
                )
            
            messages.success(request, 'Announcement updated successfully.')
            
        except Exception as e:
            messages.error(request, f'Error updating announcement: {str(e)}')
    
    return redirect('teacher_announcements')

@login_required
def teacher_delete_announcement(request, announcement_id):
    if not hasattr(request.user, 'teacher_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
    
    teacher = request.user.teacher_profile
    announcement = get_object_or_404(Announcement, id=announcement_id, created_by=request.user)
    
    if request.method == 'POST':
        try:
            # Delete associated attachments first
            announcement.announcementattachment_set.all().delete()
            
            # Delete the announcement
            announcement.delete()
            
            messages.success(request, 'Announcement deleted successfully.')
            
        except Exception as e:
            messages.error(request, f'Error deleting announcement: {str(e)}')
    
    return redirect('teacher_announcements')

@login_required
def teacher_student_detail(request, student_id):
    if not hasattr(request.user, 'teacher_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
        
    teacher = request.user.teacher_profile
    student = get_object_or_404(Student, id=student_id)
    
    # Verify that the student is enrolled in at least one of the teacher's courses
    if not student.enrolled_courses.filter(teacher=teacher).exists():
        messages.error(request, 'You do not have access to this student\'s information.')
        return redirect('teacher_students')
    
    # Get all courses the student is enrolled in with this teacher
    courses = student.enrolled_courses.filter(teacher=teacher)
    
    # Get attendance summary for each course
    attendance_summary = []
    for course in courses:
        total_classes = Attendance.objects.filter(
            student=student,
            course=course
        ).count()
        
        present_classes = Attendance.objects.filter(
            student=student,
            course=course,
            status='present'
        ).count()
        
        attendance_summary.append({
            'course': course,
            'total_classes': total_classes,
            'present_classes': present_classes,
            'percentage': (present_classes / total_classes * 100) if total_classes > 0 else 0
        })
    
    # Get recent assignments and exams
    assignments = Activity.objects.filter(
        course__in=courses,
        activity_type='assignment'
    ).order_by('-due_date')[:5]
    
    exams = Examination.objects.filter(
        course__in=courses
    ).order_by('-date')[:5]
    
    # Get recent attendance records
    recent_attendance = Attendance.objects.filter(
        student=student,
        course__in=courses
    ).order_by('-date')[:5]
    
    # Get recent grades
    recent_grades = Grade.objects.filter(
        student=student,
        subject__in=courses
    ).order_by('-date')[:5]
    
    # Calculate overall statistics
    student.attendance_percentage = student.get_attendance_percentage()
    student.performance = student.get_performance()
    
    context = {
        'student': student,
        'courses': courses,
        'attendance_summary': attendance_summary,
        'assignments': assignments,
        'exams': exams,
        'recent_attendance': recent_attendance,
        'recent_grades': recent_grades
    }
    return render(request, 'accounts/teacher/students/detail.html', context)

@login_required
def teacher_student_attendance(request, student_id):
    if not hasattr(request.user, 'teacher_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
        
    teacher = request.user.teacher_profile
    student = get_object_or_404(Student, id=student_id)
    
    # Verify that the student is enrolled in at least one of the teacher's courses
    if not student.enrolled_courses.filter(teacher=teacher).exists():
        messages.error(request, 'You do not have access to this student\'s information.')
        return redirect('teacher_students')
    
    # Get all courses the student is enrolled in with this teacher
    courses = student.enrolled_courses.filter(teacher=teacher)
    
    # Get attendance records
    attendance_records = Attendance.objects.filter(
        student=student,
        course__in=courses
    ).order_by('-date')
    
    # Calculate overall attendance statistics
    total_classes = attendance_records.count()
    total_present = attendance_records.filter(status='present').count()
    total_absent = attendance_records.filter(status='absent').count()
    attendance_percentage = (total_present / total_classes * 100) if total_classes > 0 else 0
    
    # Calculate course-wise attendance
    course_attendance = []
    for course in courses:
        course_records = attendance_records.filter(course=course)
        course_total = course_records.count()
        course_present = course_records.filter(status='present').count()
        course_absent = course_records.filter(status='absent').count()
        course_percentage = (course_present / course_total * 100) if course_total > 0 else 0
        
        course_attendance.append({
            'course': course,
            'total_classes': course_total,
            'present_count': course_present,
            'absent_count': course_absent,
            'attendance_percentage': course_percentage
        })
    
    context = {
        'student': student,
        'courses': courses,
        'attendance_records': attendance_records,
        'total_classes': total_classes,
        'total_present': total_present,
        'total_absent': total_absent,
        'attendance_percentage': attendance_percentage,
        'course_attendance': course_attendance
    }
    return render(request, 'accounts/teacher/students/attendance.html', context)

@login_required
def teacher_student_grades(request, student_id):
    if not hasattr(request.user, 'teacher_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
        
    teacher = request.user.teacher_profile
    student = get_object_or_404(Student, id=student_id)
    
    # Verify that the student is enrolled in at least one of the teacher's courses
    if not student.enrolled_courses.filter(teacher=teacher).exists():
        messages.error(request, 'You do not have access to this student\'s information.')
        return redirect('teacher_students')
    
    # Get all courses the student is enrolled in with this teacher
    all_courses = student.enrolled_courses.filter(teacher=teacher)
    selected_course_id = request.GET.get('course')
    selected_course = None
    if selected_course_id:
        try:
            selected_course = all_courses.get(id=selected_course_id)
            courses = all_courses.filter(id=selected_course_id)
        except Course.DoesNotExist:
            selected_course = None
            courses = all_courses
    else:
        courses = all_courses
    
    # Get grades summary for each course
    grades_summary = []
    for course in courses:
        # Get grades for this student in this course
        course_grades = Grade.objects.filter(
            student=student,
            subject=course
        )
        
        # Calculate total score for this course
        if course_grades.exists():
            total_score = sum(grade.score for grade in course_grades)
            max_score = course_grades.count() * 100  # Assuming grades are out of 100
        else:
            total_score = 0
            max_score = 100  # Default max score
        
        grades_summary.append({
            'course': course,
            'total_score': total_score,
            'max_score': max_score,
            'percentage': (total_score / max_score * 100) if max_score > 0 else 0
        })
    
    # Get all grades for this student in these courses with related data
    all_grades = Grade.objects.filter(
        student=student, 
        subject__in=courses
    ).select_related('subject', 'examination').order_by('-date', '-created_at')
    
    # Calculate statistics
    total_subjects = len(grades_summary)
    overall_percentage = 0
    highest_score = 0
    lowest_score = 100
    highest_grade_letter = '-'
    average_score = 0
    recent_performance_trend = []
    
    if all_grades.exists():
        scores = [float(g.score) for g in all_grades]
        total_score_sum = sum(scores)
        max_score_sum = all_grades.count() * 100
        overall_percentage = round((total_score_sum / max_score_sum * 100), 2) if max_score_sum > 0 else 0
        average_score = round((total_score_sum / len(scores)), 2) if len(scores) > 0 else 0
        highest_score = max(scores)
        lowest_score = min(scores)
        
        # Calculate automatic grades if not provided
        for grade in all_grades:
            if not grade.grade:
                if grade.score >= 90:
                    grade.calculated_grade = 'A'
                elif grade.score >= 80:
                    grade.calculated_grade = 'B'
                elif grade.score >= 70:
                    grade.calculated_grade = 'C'
                elif grade.score >= 60:
                    grade.calculated_grade = 'D'
                else:
                    grade.calculated_grade = 'F'
            else:
                grade.calculated_grade = grade.grade
        
        # Find the highest letter grade (A+ > A > B > C > D > F)
        grade_order = ['A+', 'A', 'B', 'C', 'D', 'F']
        grades_letters = [g.grade or g.calculated_grade for g in all_grades if hasattr(g, 'calculated_grade')]
        for letter in grade_order:
            if letter in grades_letters:
                highest_grade_letter = letter
                break
        
        # Recent performance trend (last 5 grades)
        recent_grades = list(all_grades[:5])
        recent_performance_trend = [float(g.score) for g in recent_grades]
    context = {
        'student': student,
        'courses': all_courses,
        'selected_course': selected_course,
        'grades_summary': grades_summary,
        'all_grades': all_grades,
        'overall_percentage': overall_percentage,
        'total_subjects': total_subjects,
        'highest_grade': highest_grade_letter,
        'highest_score': highest_score,
        'lowest_score': lowest_score,
        'average_score': average_score,
        'recent_performance_trend': recent_performance_trend,
        'total_grades': all_grades.count(),
        'today': timezone.now().date()
    }
    return render(request, 'accounts/teacher/students/grades.html', context)

@login_required
def teacher_export_students(request):
    if not hasattr(request.user, 'teacher_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
        
    if request.method == 'POST':
        teacher = request.user.teacher_profile
        format = request.POST.get('format', 'csv')
        fields = request.POST.getlist('fields', [])
        
        # Get students data based on selected fields
        students = Student.objects.filter(enrolled_courses__teacher=teacher).distinct()
        
        # Create response based on format
        if format == 'csv':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="students.csv"'
            
            writer = csv.writer(response)
            # Write header
            writer.writerow(['Student ID', 'Name', 'Email', 'Phone', 'Course', 'Class'])
            
            # Write data
            for student in students:
                writer.writerow([
                    student.student_id,
                    student.get_full_name(),
                    student.email,
                    student.phone_number,
                    student.course.name,
                    student.class_section.name
                ])
            
            return response
            
        elif format == 'excel':
            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename="students.xlsx"'
            
            workbook = xlsxwriter.Workbook(response)
            worksheet = workbook.add_worksheet()
            
            # Write header
            headers = ['Student ID', 'Name', 'Email', 'Phone', 'Course', 'Class']
            for col, header in enumerate(headers):
                worksheet.write(0, col, header)
            
            # Write data
            for row, student in enumerate(students, start=1):
                worksheet.write(row, 0, student.student_id)
                worksheet.write(row, 1, student.get_full_name())
                worksheet.write(row, 2, student.email)
                worksheet.write(row, 3, student.phone_number)
                worksheet.write(row, 4, student.course.name)
                worksheet.write(row, 5, student.class_section.name)
            
            workbook.close()
            return response
            
        elif format == 'pdf':
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="students.pdf"'
            
            # Create PDF
            p = canvas.Canvas(response)
            p.drawString(100, 800, "Students List")
            
            y = 750
            for student in students:
                p.drawString(100, y, f"{student.student_id} - {student.get_full_name()}")
                y -= 20
                if y < 50:
                    p.showPage()
                    y = 750
            
            p.save()
            return response
    
    return redirect('teacher_students')

@login_required
def teacher_exam_edit(request, exam_id):
    if not hasattr(request.user, 'teacher_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
        
    teacher = request.user.teacher_profile
    exam = get_object_or_404(Examination, id=exam_id, course__teacher=teacher)
    
    if request.method == 'POST':
        course_id = request.POST.get('course')
        title = request.POST.get('title')
        exam_type = request.POST.get('exam_type')
        date_str = request.POST.get('date')
        duration = request.POST.get('duration')
        total_marks = request.POST.get('total_marks')
        description = request.POST.get('description')

        from datetime import datetime, timedelta
        try:
            course = Course.objects.get(id=course_id, teacher=teacher)
            exam.course = course
            exam.title = title
            exam.exam_type = exam_type
            # Parse date string to datetime object
            if date_str:
                exam.date = datetime.strptime(date_str, '%Y-%m-%dT%H:%M')
            # Handle duration as minutes (if provided)
            if duration:
                exam.duration = timedelta(minutes=int(duration))
            exam.total_marks = total_marks
            exam.instructions = description
            exam.save()

            messages.success(request, 'Exam updated successfully.')
            return redirect('teacher_exam_detail', exam_id=exam.id)
        except Course.DoesNotExist:
            messages.error(request, 'Invalid course selected.')
        except Exception as e:
            messages.error(request, f'Error updating exam: {str(e)}')
    
    courses = Course.objects.filter(teacher=teacher)
    exam_type_choices = Examination.EXAM_TYPES
    context = {
        'exam': exam,
        'courses': courses,
        'exam_type_choices': exam_type_choices
    }
    return render(request, 'accounts/teacher/exams/edit.html', context)

@login_required
def teacher_exam_delete(request, exam_id):
    if not hasattr(request.user, 'teacher_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
        
    teacher = request.user.teacher_profile
    exam = get_object_or_404(Examination, id=exam_id, course__teacher=teacher)
    
    if request.method == 'POST':
        try:
            exam.delete()
            messages.success(request, 'Exam deleted successfully.')
            return redirect('teacher_exams')
        except Exception as e:
            messages.error(request, f'Error deleting exam: {str(e)}')
            return redirect('teacher_exam_detail', exam_id=exam.id)
    
    context = {
        'exam': exam
    }
    return render(request, 'accounts/teacher/exams/delete.html', context)

@login_required
def student_timetable_data(request):
    if not hasattr(request.user, 'student_profile'):
        return JsonResponse({'error': 'Access denied'}, status=403)
        
    student = request.user.student_profile
    date_param = request.GET.get('date')
    
    try:
        selected_date = datetime.strptime(date_param, '%Y-%m-%dT%H:%M:%S.%fZ').date() if date_param else datetime.now().date()
    except ValueError:
        selected_date = datetime.now().date()
    
    # Get the week's start and end dates
    week_start = selected_date - timedelta(days=selected_date.weekday())
    week_end = week_start + timedelta(days=6)
    
    # Get all timetable entries for the student's courses
    timetable_entries = Timetable.objects.filter(
        course__in=student.enrolled_courses.all()
    ).order_by('day', 'start_time')
    
    # Format the data for the frontend
    timetable_data = []
    for entry in timetable_entries:
        timetable_data.append({
            'id': entry.id,
            'course': entry.course.title,
            'day': entry.day,
            'start_time': entry.start_time.strftime('%H:%M'),
            'end_time': entry.end_time.strftime('%H:%M'),
            'room': entry.room,
            'teacher': entry.course.teacher.get_full_name()
        })
    
    return JsonResponse({
        'week_start': week_start.strftime('%Y-%m-%d'),
        'week_end': week_end.strftime('%Y-%m-%d'),
        'timetable': timetable_data
    })

@login_required
def teacher_create_course(request):
    if not hasattr(request.user, 'teacher_profile'):
        return JsonResponse({'success': False, 'message': 'You do not have access to this feature.'})
        
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method.'})
    
    teacher = request.user.teacher_profile
    
    try:
        course_code = request.POST.get('course_code')
        title = request.POST.get('title')
        description = request.POST.get('description')
        credits = request.POST.get('credits')
        class_section_id = request.POST.get('class_section')
        
        # Validate required fields
        if not all([course_code, title, description, credits, class_section_id]):
            return JsonResponse({'success': False, 'message': 'All fields are required.'})
            
        # Check if course code already exists
        if Course.objects.filter(course_code=course_code).exists():
            return JsonResponse({'success': False, 'message': 'Course code already exists.'})
            
        # Get the class section
        try:
            class_section = Class.objects.get(id=class_section_id)
        except Class.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Invalid class section.'})
            
        # Create the course
        course = Course.objects.create(
            course_code=course_code,
            title=title,
            description=description,
            credits=int(credits),
            teacher=teacher,
            class_section=class_section
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Course created successfully.',
            'course': {
                'id': course.id,
                'code': course.course_code,
                'title': course.title
            }
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required
def teacher_create_class(request):
    if not hasattr(request.user, 'teacher_profile'):
        return JsonResponse({'success': False, 'message': 'You do not have access to this feature.'})
        
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method.'})
    
    print('Received POST request for class creation')
    print('POST data:', request.POST)
    
    try:
        name = request.POST.get('name')
        section = request.POST.get('section')
        academic_year = request.POST.get('academic_year')
        
        print('Extracted data:', {
            'name': name,
            'section': section,
            'academic_year': academic_year
        })
        
        # Validate required fields
        if not all([name, section, academic_year]):
            return JsonResponse({'success': False, 'message': 'All fields are required.'})
            
        # Check if class with same name, section and academic year exists
        if Class.objects.filter(name=name, section=section, academic_year=academic_year).exists():
            return JsonResponse({'success': False, 'message': 'Class with these details already exists.'})
            
        # Create the class
        class_obj = Class.objects.create(
            name=name,
            section=section,
            academic_year=academic_year
        )
        
        print('Class created successfully:', class_obj)
        
        return JsonResponse({
            'success': True,
            'message': 'Class created successfully.',
            'class': {
                'id': class_obj.id,
                'name': class_obj.name,
                'section': class_obj.section,
                'academic_year': class_obj.academic_year
            }
        })
        
    except Exception as e:
        print('Error creating class:', str(e))
        return JsonResponse({'success': False, 'message': str(e)})

@login_required
def teacher_course_detail_view(request, id):
    if not hasattr(request.user, 'teacher_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
        
    teacher = request.user.teacher_profile
    course = get_object_or_404(Course, id=id, teacher=teacher)
    
    # Get course statistics
    total_students = course.students.count()
    total_assignments = Activity.objects.filter(course=course, activity_type='assignment').count()
    total_exams = Examination.objects.filter(course=course).count()
    
    # Get attendance statistics
    total_classes = Attendance.objects.filter(course=course).values('date').distinct().count()
    attendance_records = Attendance.objects.filter(course=course)
    attendance_rate = (
        attendance_records.filter(status='present').count() / attendance_records.count() * 100
        if attendance_records.count() > 0 else 0
    )
    
    context = {
        'course': course,
        'total_students': total_students,
        'total_assignments': total_assignments,
        'total_exams': total_exams,
        'total_classes': total_classes,
        'attendance_rate': round(attendance_rate, 2),
        'students': course.students.all(),
        'recent_activities': Activity.objects.filter(course=course).order_by('-created_at')[:5],
        'upcoming_exams': Examination.objects.filter(course=course, date__gte=datetime.now()).order_by('date')[:3]
    }
    return render(request, 'accounts/teacher/courses/detail.html', context)

@login_required
def teacher_add_schedule_class(request):
    if not hasattr(request.user, 'teacher_profile'):
        return JsonResponse({'success': False, 'message': 'You do not have access to this feature.'})
        
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method.'})
    
    try:
        course_id = request.POST.get('course')
        day = request.POST.get('day')
        start_time = request.POST.get('start_time')
        end_time = request.POST.get('end_time')
        room = request.POST.get('room')
        
        # Validate required fields
        if not all([course_id, day, start_time, end_time, room]):
            return JsonResponse({'success': False, 'message': 'All fields are required.'})
            
        # Get the course and verify ownership
        course = get_object_or_404(Course, id=course_id)
        if course.teacher != request.user.teacher_profile:
            return JsonResponse({'success': False, 'message': 'You do not have permission to schedule this course.'})
            
        # Check for schedule conflicts
        existing_schedule = Schedule.objects.filter(
            course__teacher=request.user.teacher_profile,
            day=day,
            start_time__lt=end_time,
            end_time__gt=start_time
        ).exists()
        
        if existing_schedule:
            return JsonResponse({'success': False, 'message': 'There is a scheduling conflict with another class.'})
            
        # Create the schedule
        Schedule.objects.create(
            course=course,
            day=day,
            start_time=start_time,
            end_time=end_time,
            room=room
        )
        
        return JsonResponse({'success': True, 'message': 'Class has been added to schedule successfully.'})
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required
def teacher_compose_message(request):
    if not hasattr(request.user, 'teacher_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
        
    teacher = request.user.teacher_profile
    
    # Get all students from teacher's courses
    students = Student.objects.filter(enrolled_courses__teacher=teacher).distinct()
    
    # Get all teachers except current user
    teachers = Teacher.objects.exclude(id=teacher.id)
    
    # Get all classes and courses for the teacher
    classes = Class.objects.filter(courses__teacher=teacher).distinct()
    courses = Course.objects.filter(teacher=teacher)
    
    context = {
        'students': students,
        'teachers': teachers,
        'classes': classes,
        'courses': courses
    }
    return render(request, 'accounts/teacher/communication/compose_message.html', context)

@login_required
def teacher_send_message(request):
    if not hasattr(request.user, 'teacher_profile') or request.method != 'POST':
        messages.error(request, 'Invalid request.')
        return redirect('teacher_messages')
    
    teacher = request.user.teacher_profile
    recipient_type = request.POST.get('recipient_type')
    subject = request.POST.get('subject')
    content = request.POST.get('content')
    attachment = request.FILES.get('attachment')
    
    try:
        recipients = []
        
        if recipient_type == 'reply':
            recipient_id = request.POST.get('recipient_id')
            if recipient_id:
                recipients.append(User.objects.get(id=recipient_id))
        elif recipient_type == 'student':
            student_user_id = request.POST.get('recipient_student')
            if student_user_id:
                recipients.append(User.objects.get(id=student_user_id))
                
        elif recipient_type == 'teacher':
            teacher_user_id = request.POST.get('recipient_teacher')
            if teacher_user_id:
                recipients.append(User.objects.get(id=teacher_user_id))
                
        elif recipient_type == 'class':
            class_id = request.POST.get('recipient_class')
            if class_id:
                class_obj = Class.objects.get(id=class_id)
                recipients.extend([student.user for student in class_obj.students.all()])
                
        elif recipient_type == 'course':
            course_id = request.POST.get('recipient_course')
            if course_id:
                course = Course.objects.get(id=course_id)
                recipients.extend([student.user for student in course.students.all()])
        
        if not recipients:
            messages.error(request, 'Please select at least one recipient.')
            return redirect('teacher_compose_message')
        
        # Create message for each recipient
        for recipient in recipients:
            message = Message.objects.create(
                sender=request.user,
                recipient=recipient,
                subject=subject,
                content=content
            )
            
            # Handle attachment if provided
            if attachment:
                MessageAttachment.objects.create(
                    message=message,
                    file=attachment,
                    filename=attachment.name,
                    file_size=attachment.size
                )
        
        messages.success(request, f'Message sent successfully to {len(recipients)} recipient(s).')
        return redirect('teacher_messages')
        
    except Exception as e:
        messages.error(request, f'Error sending message: {str(e)}')
        return redirect('teacher_compose_message')

@login_required
def view_payment_details(request, payment_id):
    try:
        payment = Payment.objects.get(id=payment_id)
        context = {
            'payment': payment,
            'page_title': 'Payment Details'
        }
        return render(request, 'accounts/payment_details.html', context)
    except Payment.DoesNotExist:
        messages.error(request, 'Payment not found.')
        return redirect('fee_payment')

@login_required
def print_payment_receipt(request, payment_id):
    try:
        payment = Payment.objects.get(id=payment_id)
        context = {
            'payment': payment,
            'page_title': 'Payment Receipt'
        }
        return render(request, 'accounts/payment_receipt.html', context)
    except Payment.DoesNotExist:
        messages.error(request, 'Payment not found.')
        return redirect('fee_payment')

@login_required
def view_invoice_details(request, invoice_id):
    try:
        invoice = Invoice.objects.get(id=invoice_id)
        context = {
            'invoice': invoice,
            'page_title': 'Invoice Details'
        }
        return render(request, 'accounts/invoice_details.html', context)
    except Invoice.DoesNotExist:
        messages.error(request, 'Invoice not found.')
        return redirect('invoices')

@login_required
def print_invoice(request, invoice_id):
    try:
        invoice = Invoice.objects.get(id=invoice_id)
        context = {
            'invoice': invoice,
            'page_title': 'Invoice'
        }
        return render(request, 'accounts/invoice_print.html', context)
    except Invoice.DoesNotExist:
        messages.error(request, 'Invoice not found.')
        return redirect('invoices')

@login_required
def delete_invoice(request, invoice_id):
    invoice = get_object_or_404(Invoice, id=invoice_id)
    if not request.user.is_staff:
        messages.error(request, 'You do not have permission to delete invoices.')
        return redirect('invoices')
    if request.method == 'POST':
        invoice.delete()
        messages.success(request, 'Invoice deleted successfully.')
        return redirect('invoices')
    context = {'invoice': invoice, 'page_title': 'Delete Invoice'}
    return render(request, 'accounts/invoice_confirm_delete.html', context)

@login_required
def generate_student_report(request):
    if request.method == 'POST':
        student_id = request.POST.get('student')
        report_type = request.POST.get('report_type')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        
        # Get student data
        student = Student.objects.get(id=student_id)
        
        # Generate report based on type
        if report_type == 'academic':
            # Get all courses the student is enrolled in
            courses = student.enrolled_courses.all()
            
            # Get academic performance data
            grades = Grade.objects.filter(
                student=student,
                date__range=[start_date, end_date]
            ).select_related('subject')
            
            # Calculate overall statistics
            total_score = 0
            total_grades = 0
            highest_score = 0
            lowest_score = 100
            
            # Calculate subject-wise performance
            subject_performance = []
            for course in courses:
                course_grades = grades.filter(subject=course)
                if course_grades.exists():
                    course_total = sum(g.score for g in course_grades)
                    course_count = course_grades.count()
                    course_avg = course_total / course_count
                    course_max = max(g.score for g in course_grades)
                    course_min = min(g.score for g in course_grades)
                    
                    # Update overall statistics
                    total_score += course_total
                    total_grades += course_count
                    highest_score = max(highest_score, course_max)
                    lowest_score = min(lowest_score, course_min)
                    
                    subject_performance.append({
                        'subject': course,
                        'average': round(course_avg, 1),
                        'highest': course_max,
                        'lowest': course_min,
                        'grades': course_grades
                    })
            
            # Calculate overall average
            overall_average = round(total_score / total_grades, 1) if total_grades > 0 else 0
            
            # Get attendance data for the period
            attendance_records = Attendance.objects.filter(
                student=student,
                date__range=[start_date, end_date]
            )
            attendance_rate = (
                attendance_records.filter(status='present').count() / attendance_records.count() * 100
            ) if attendance_records.exists() else 0
            
            context = {
                'student': student,
                'subject_performance': subject_performance,
                'overall_average': overall_average,
                'highest_score': highest_score,
                'lowest_score': lowest_score,
                'attendance_rate': round(attendance_rate, 1),
                'total_subjects': len(subject_performance),
                'total_grades': total_grades,
                'report_type': 'Academic Performance',
                'start_date': start_date,
                'end_date': end_date
            }
            template = 'accounts/reports/student_academic.html'
            
        elif report_type == 'attendance':
            # Get attendance data
            attendance = Attendance.objects.filter(
                student=student,
                date__range=[start_date, end_date]
            )
            context = {
                'student': student,
                'attendance': attendance,
                'report_type': 'Attendance',
                'start_date': start_date,
                'end_date': end_date
            }
            template = 'accounts/reports/student_attendance.html'
            
        else:  # behavior
            # Get behavior data
            behavior = Behavior.objects.filter(
                student=student,
                date__range=[start_date, end_date]
            )
            context = {
                'student': student,
                'behavior': behavior,
                'report_type': 'Behavior',
                'start_date': start_date,
                'end_date': end_date
            }
            template = 'accounts/reports/student_behavior.html'
        
        return render(request, template, context)
    
    return redirect('reports')

@login_required
def generate_teacher_report(request):
    if request.method == 'POST':
        teacher_id = request.POST.get('teacher')
        report_type = request.POST.get('report_type')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        
        # Get teacher data
        teacher = Teacher.objects.get(id=teacher_id)
        
        # Generate report based on type
        if report_type == 'teaching':
            # Get teaching hours data from schedules
            courses = Course.objects.filter(teacher=teacher)
            schedules = Schedule.objects.filter(course__in=courses)
            
            # Calculate total hours per week
            total_hours = 0
            for schedule in schedules:
                start = datetime.strptime(str(schedule.start_time), '%H:%M:%S')
                end = datetime.strptime(str(schedule.end_time), '%H:%M:%S')
                duration = end - start
                total_hours += duration.total_seconds() / 3600  # Convert to hours
            
            teaching_hours = {
                'total_hours': round(total_hours, 2),
                'classes_taught': courses.count(),
                'avg_hours_per_day': round(total_hours / 5, 2) if courses.exists() else 0
            }
            context = {
                'teacher': teacher,
                'teaching_hours': teaching_hours,
                'report_type': 'Teaching Hours',
                'start_date': start_date,
                'end_date': end_date
            }
            template = 'accounts/reports/teacher_teaching.html'
            
        elif report_type == 'performance':
            # Get subject performance data
            courses = Course.objects.filter(teacher=teacher)
            subject_performance = []
            
            # Overall statistics
            total_students = 0
            overall_pass_rate = 0
            overall_avg_score = 0
            courses_with_data = 0
            
            for course in courses:
                # Get all grades for this course in the date range
                grades = Grade.objects.filter(
                    subject=course,
                    date__range=[start_date, end_date]
                ).select_related('student')
                
                if grades.exists():
                    # Calculate statistics
                    total_score = sum(grade.score for grade in grades)
                    avg_score = total_score / grades.count()
                    pass_count = sum(1 for grade in grades if grade.score >= 60)
                    pass_rate = (pass_count / grades.count() * 100)
                    top_score = max(grade.score for grade in grades)
                    
                    # Count unique students
                    unique_students = len(set(grade.student.id for grade in grades))
                    total_students += unique_students
                    
                    # Update overall statistics
                    overall_pass_rate += pass_rate
                    overall_avg_score += avg_score
                    courses_with_data += 1
                    
                    # Add to subject performance list
                    subject_performance.append({
                        'subject': {
                            'title': course.title,
                            'id': course.id
                        },
                        'class_name': course.class_section.name if course.class_section else 'N/A',
                        'average_score': round(avg_score, 1),
                        'pass_rate': round(pass_rate, 1),
                        'top_score': round(top_score, 1),
                        'student_count': unique_students
                    })
            
            # Calculate overall statistics
            if courses_with_data > 0:
                overall_pass_rate = round(overall_pass_rate / courses_with_data, 1)
                overall_avg_score = round(overall_avg_score / courses_with_data, 1)
            
            context = {
                'teacher': teacher,
                'subject_performance': subject_performance,  # Pass as list for template
                'overall_stats': {
                    'total_subjects': courses.count(),
                    'total_students': total_students,
                    'overall_pass_rate': overall_pass_rate,
                    'overall_avg_score': overall_avg_score
                },
                'report_type': 'Subject Performance',
                'start_date': start_date,
                'end_date': end_date
            }
            template = 'accounts/reports/teacher_performance.html'
            
        else:  # evaluation
            # Get student evaluations
            evaluations = Evaluation.objects.filter(
                teacher=teacher,
                date__range=[start_date, end_date]
            )
            
            # Serialize evaluations for JavaScript
            evaluations_data = []
            for eval in evaluations:
                evaluations_data.append({
                    'date': eval.date.strftime('%Y-%m-%d'),
                    'course': {
                        'title': eval.course.title,
                        'id': eval.course.id
                    },
                    'student': eval.student.get_full_name(),
                    'rating': eval.rating,
                    'get_rating_display': eval.get_rating_display(),
                    'comments': eval.comments
                })
            
            context = {
                'teacher': teacher,
                'evaluations': evaluations,
                'evaluations_json': json.dumps(evaluations_data),
                'report_type': 'Student Evaluations',
                'start_date': start_date,
                'end_date': end_date
            }
            template = 'accounts/reports/teacher_evaluation.html'
        
        return render(request, template, context)
    
    return redirect('reports')

@login_required
def generate_academic_report(request):
    if request.method == 'POST':
        class_id = request.POST.get('class')
        report_type = request.POST.get('report_type')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        
        # Get class data
        class_obj = Class.objects.get(id=class_id)
        
        # Generate report based on type
        if report_type == 'performance':
            # Get class performance data
            students = Student.objects.filter(class_section=class_obj)
            performance_data = []
            total_grades = 0
            total_score = 0
            grade_counts = {'A': 0, 'B': 0, 'C': 0, 'D': 0, 'F': 0}
            
            # Get all subjects for this class
            subjects = Course.objects.filter(class_section=class_obj)
            
            # Calculate attendance rate
            attendance_records = Attendance.objects.filter(
                student__in=students,
                date__range=[start_date, end_date]
            )
            total_attendance = attendance_records.count()
            present_attendance = attendance_records.filter(status='present').count()
            attendance_rate = (present_attendance / total_attendance * 100) if total_attendance > 0 else 0
            
            for student in students:
                # Get grades for all subjects
                student_grades = []
                student_total = 0
                grades_count = 0
                
                for subject in subjects:
                    grade = Grade.objects.filter(
                        student=student,
                        subject=subject,
                        date__range=[start_date, end_date]
                    ).order_by('-date').first()
                    
                    if grade:
                        student_grades.append(grade)
                        student_total += grade.score
                        grades_count += 1
                        
                        # Update grade distribution
                        if grade.score >= 90:
                            grade_counts['A'] += 1
                        elif grade.score >= 80:
                            grade_counts['B'] += 1
                        elif grade.score >= 70:
                            grade_counts['C'] += 1
                        elif grade.score >= 60:
                            grade_counts['D'] += 1
                        else:
                            grade_counts['F'] += 1
                
                # Calculate student's average
                student_avg = student_total / grades_count if grades_count > 0 else 0
                total_score += student_total
                total_grades += grades_count
                
                # Calculate student's attendance
                student_attendance = attendance_records.filter(student=student)
                student_attendance_rate = (
                    student_attendance.filter(status='present').count() / student_attendance.count() * 100
                ) if student_attendance.exists() else 0
                
                performance_data.append({
                    'student': student,
                    'grades': student_grades,
                    'average': student_avg,
                    'attendance': student_attendance_rate
                })
            
            # Calculate overall class average
            class_average = total_score / total_grades if total_grades > 0 else 0
            
            # Calculate pass rate
            total_grades = sum(grade_counts.values())
            passing_grades = sum(grade_counts[grade] for grade in ['A', 'B', 'C', 'D'])
            pass_rate = (passing_grades / total_grades * 100) if total_grades > 0 else 0
            
            # Prepare grade distribution for chart
            grade_distribution = [grade_counts[grade] for grade in ['A', 'B', 'C', 'D', 'F']]
            
            # Calculate subject averages
            subject_names = []
            subject_averages = []
            for subject in subjects:
                subject_grades = Grade.objects.filter(
                    subject=subject,
                    student__in=students,
                    date__range=[start_date, end_date]
                )
                if subject_grades.exists():
                    avg = subject_grades.aggregate(Avg('score'))['score__avg'] or 0
                    subject_names.append(f"'{subject.title}'")
                    subject_averages.append(round(avg, 1))
            
            context = {
                'class': class_obj,
                'performance_data': performance_data,
                'subjects': subjects,
                'average_grade': round(class_average, 1),
                'pass_rate': round(pass_rate, 1),
                'attendance_rate': round(attendance_rate, 1),
                'grade_distribution': grade_distribution,
                'subject_names': '[' + ', '.join(subject_names) + ']',
                'subject_averages': subject_averages,
                'report_type': 'Class Performance',
                'start_date': start_date,
                'end_date': end_date
            }
            template = 'accounts/reports/academic_performance.html'
            
        elif report_type == 'exam':
            # Get exam results
            exams = Examination.objects.filter(
                course__class_section=class_obj,
                date__range=[start_date, end_date]
            )
            context = {
                'class': class_obj,
                'exams': exams,
                'report_type': 'Exam Results',
                'start_date': start_date,
                'end_date': end_date
            }
            template = 'accounts/reports/academic_exam.html'
            
        else:  # grade
            # Get grade analysis
            grades = Grade.objects.filter(
                student__class_section=class_obj,
                date__range=[start_date, end_date]
            )
            context = {
                'class': class_obj,
                'grades': grades,
                'report_type': 'Grade Analysis',
                'start_date': start_date,
                'end_date': end_date
            }
            template = 'accounts/reports/academic_grade.html'
        
        return render(request, template, context)
    
    return redirect('reports')

@login_required
def generate_financial_report(request):
    if request.method == 'POST':
        report_type = request.POST.get('report_type')
        department_id = request.POST.get('department')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        
        # Generate report based on type
        if report_type == 'fees':
            # Get fee collection data
            fees = Fee.objects.filter(
                created_at__range=[start_date, end_date]
            )
            if department_id:
                fees = fees.filter(department_id=department_id)
            
            # Calculate summary statistics
            total_amount = fees.aggregate(total=Sum('amount'))['total'] or 0
            total_transactions = fees.count()
            average_amount = total_amount / total_transactions if total_transactions > 0 else 0
            
            context = {
                'fees': fees,
                'total_amount': total_amount,
                'total_transactions': total_transactions,
                'average_amount': average_amount,
                'report_type': 'Fee Collection',
                'start_date': start_date,
                'end_date': end_date
            }
            template = 'accounts/reports/financial_fees.html'
            
        elif report_type == 'expenses':
            # Get expenses data
            expenses = Expense.objects.filter(
                created_at__range=[start_date, end_date]
            )
            if department_id:
                expenses = expenses.filter(department_id=department_id)
            context = {
                'expenses': expenses,
                'report_type': 'Expenses',
                'start_date': start_date,
                'end_date': end_date
            }
            template = 'accounts/reports/financial_expenses.html'
            
        else:  # summary
            # Get financial summary
            fees = Fee.objects.filter(
                created_at__range=[start_date, end_date]
            )
            expenses = Expense.objects.filter(
                created_at__range=[start_date, end_date]
            )
            if department_id:
                fees = fees.filter(department_id=department_id)
                expenses = expenses.filter(department_id=department_id)
            context = {
                'fees': fees,
                'expenses': expenses,
                'report_type': 'Financial Summary',
                'start_date': start_date,
                'end_date': end_date
            }
            template = 'accounts/reports/financial_summary.html'
        
        return render(request, template, context)
    
    return redirect('reports')

@login_required
def student_academic_update(request):
    if not hasattr(request.user, 'student_profile'):
        messages.error(request, 'You do not have a student profile.')
        return redirect('student_dashboard')
    
    if request.method == 'POST':
        student = request.user.student_profile
        try:
            # Get form data
            batch = request.POST.get('batch')
            admission_date = request.POST.get('admission_date')
            grade_level = request.POST.get('grade_level')
            roll_number = request.POST.get('roll_number')
            academic_status = request.POST.get('academic_status')

            # Validate required fields
            if not all([grade_level, admission_date]):
                messages.error(request, 'Grade level and admission date are required fields.')
                return redirect('student_profile')

            # Update student information
            student.batch = batch
            student.admission_date = admission_date
            student.grade_level = grade_level
            student.roll_number = roll_number
            student.status = academic_status if academic_status else 'active'
            
            student.save()
            messages.success(request, 'Academic information updated successfully.')
        except Exception as e:
            messages.error(request, f'Error updating academic information: {str(e)}')
    
    return redirect('student_profile')

@login_required
def teacher_edit_course(request, course_id):
    if not hasattr(request.user, 'teacher_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
        
    teacher = request.user.teacher_profile
    course = get_object_or_404(Course, id=course_id, teacher=teacher)
    
    if request.method == 'POST':
        try:
            course.course_code = request.POST.get('course_code')
            course.title = request.POST.get('title')
            course.description = request.POST.get('description')
            course.credits = request.POST.get('credits')
            class_section_id = request.POST.get('class_section')
            
            # Validate required fields
            if not all([course.course_code, course.title, course.description, course.credits, class_section_id]):
                messages.error(request, 'All fields are required.')
                return redirect('teacher_edit_course', course_id=course.id)
                
            # Check if course code already exists (excluding current course)
            if Course.objects.filter(course_code=course.course_code).exclude(id=course.id).exists():
                messages.error(request, 'Course code already exists.')
                return redirect('teacher_edit_course', course_id=course.id)
                
            # Get the class section
            try:
                class_section = Class.objects.get(id=class_section_id)
                course.class_section = class_section
            except Class.DoesNotExist:
                messages.error(request, 'Invalid class section.')
                return redirect('teacher_edit_course', course_id=course.id)
                
            course.save()
            messages.success(request, 'Course updated successfully.')
            return redirect('teacher_courses')
            
        except Exception as e:
            messages.error(request, f'Error updating course: {str(e)}')
            return redirect('teacher_edit_course', course_id=course.id)
    
    # Get all available classes
    classes = Class.objects.all().order_by('name', 'section')
    
    context = {
        'course': course,
        'classes': classes
    }
    return render(request, 'accounts/teacher/courses/edit.html', context)

@login_required
def teacher_delete_course(request, course_id):
    if not hasattr(request.user, 'teacher_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
        
    teacher = request.user.teacher_profile
    course = get_object_or_404(Course, id=course_id, teacher=teacher)
    
    if request.method == 'POST':
        try:
            course.delete()
            messages.success(request, 'Course deleted successfully.')
            return redirect('teacher_courses')
        except Exception as e:
            messages.error(request, f'Error deleting course: {str(e)}')
            return redirect('teacher_courses')
    
    context = {
        'course': course
    }
    return render(request, 'accounts/teacher/courses/delete.html', context)

@login_required
def teacher_update_grades(request, course_id, student_id):
    if not hasattr(request.user, 'teacher_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
    
    teacher = request.user.teacher_profile
    course = get_object_or_404(Course, id=course_id, teacher=teacher)
    student = get_object_or_404(Student, id=student_id)
    
    if request.method == 'POST':
        # Add grade creation logic
        subject_id = request.POST.get('subject')
        score = request.POST.get('score')
        grade_letter = request.POST.get('grade')
        remarks = request.POST.get('remarks')
        date = request.POST.get('date')
        if subject_id and score and grade_letter and date:
            try:
                subject = Course.objects.get(id=subject_id, teacher=teacher)
                Grade.objects.create(
                    student=student,
                    subject=subject,
                    score=score,
                    grade=grade_letter,
                    remarks=remarks,
                    date=date
                )
                messages.success(request, 'Grade added successfully.')
                return redirect('teacher_student_grades', student_id=student.id)
            except Exception as e:
                messages.error(request, f'Error adding grade: {str(e)}')
                return redirect('teacher_student_grades', student_id=student.id)
        # Existing logic for updating assignments/exams (if any)
        try:
            for key, value in request.POST.items():
                if key.startswith('assignment_'):
                    activity_id = key.split('_')[1]
                    activity = Activity.objects.get(id=activity_id, course=course)
                    Submission.objects.update_or_create(
                        student=student,
                        activity=activity,
                        defaults={'score': value}
                    )
                elif key.startswith('exam_'):
                    exam_id = key.split('_')[1]
                    exam = Examination.objects.get(id=exam_id, course=course)
                    Report.objects.update_or_create(
                        student=student,
                        examination=exam,
                        defaults={'marks': value}
                    )
            messages.success(request, 'Grades updated successfully.')
        except Exception as e:
            messages.error(request, f'Error updating grades: {str(e)}')
        return redirect('teacher_student_grades', student_id=student.id)
    return redirect('teacher_student_grades', student_id=student.id)

@login_required
def teacher_edit_grade(request, grade_id, student_id):
    if not hasattr(request.user, 'teacher_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
    
    teacher = request.user.teacher_profile
    student = get_object_or_404(Student, id=student_id)
    grade = get_object_or_404(Grade, id=grade_id, student=student, subject__teacher=teacher)
    
    if request.method == 'POST':
        try:
            score = request.POST.get('score')
            grade_letter = request.POST.get('grade')
            remarks = request.POST.get('remarks')
            date = request.POST.get('date')
            
            if score and grade_letter and date:
                grade.score = float(score)
                grade.grade = grade_letter
                grade.remarks = remarks
                grade.date = date
                grade.save()
                
                messages.success(request, f'Grade updated successfully for {student.get_full_name()}.')
            else:
                messages.error(request, 'Please fill in all required fields.')
                
        except Exception as e:
            messages.error(request, f'Error updating grade: {str(e)}')
    
    return redirect('teacher_student_grades', student_id=student.id)

@login_required
def teacher_delete_grade(request, grade_id, student_id):
    if not hasattr(request.user, 'teacher_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
    
    teacher = request.user.teacher_profile
    student = get_object_or_404(Student, id=student_id)
    grade = get_object_or_404(Grade, id=grade_id, student=student, subject__teacher=teacher)
    
    if request.method == 'POST':
        try:
            subject_title = grade.subject.title
            grade_score = grade.score
            grade.delete()
            messages.success(request, f'Grade for {subject_title} (Score: {grade_score}) has been deleted successfully.')
        except Exception as e:
            messages.error(request, f'Error deleting grade: {str(e)}')
    
    return redirect('teacher_student_grades', student_id=student.id)

@login_required
def teacher_student_grades_data(request, student_id):
    """API endpoint to get student grade data for AJAX requests"""
    if not hasattr(request.user, 'teacher_profile'):
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    teacher = request.user.teacher_profile
    student = get_object_or_404(Student, id=student_id)
    
    # Get all grades for this student in courses taught by this teacher
    grades = Grade.objects.filter(
        student=student, 
        subject__teacher=teacher
    ).select_related('subject').order_by('-date')
    
    grades_data = []
    for grade in grades:
        grades_data.append({
            'id': grade.id,
            'subject': grade.subject.title,
            'score': float(grade.score),
            'grade': grade.grade or '',
            'date': grade.date.strftime('%Y-%m-%d'),
            'remarks': grade.remarks or ''
        })
    
    return JsonResponse({'grades': grades_data})

@login_required
def teacher_profile_edit(request):
    if not hasattr(request.user, 'teacher_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
    
    teacher = request.user.teacher_profile
    
    if request.method == 'POST':
        try:
            # Handle profile picture update
            if 'photo' in request.FILES:
                teacher.photo = request.FILES['photo']
            
            # Update basic information
            teacher.first_name = request.POST.get('first_name')
            teacher.last_name = request.POST.get('last_name')
            teacher.user.first_name = request.POST.get('first_name')
            teacher.user.last_name = request.POST.get('last_name')
            teacher.user.email = request.POST.get('email')
            
            # Update professional information
            teacher.designation = request.POST.get('designation')
            teacher.qualification = request.POST.get('qualification')
            teacher.specialization = request.POST.get('specialization')
            if request.POST.get('experience'):
                teacher.experience = int(request.POST.get('experience'))
            
            # Update contact information
            teacher.phone_number = request.POST.get('phone_number')  # Changed from phone to phone_number
            teacher.emergency_contact = request.POST.get('emergency_contact')
            teacher.address = request.POST.get('address')
            teacher.city = request.POST.get('city')
            teacher.state = request.POST.get('state')
            teacher.country = request.POST.get('country')
            teacher.postal_code = request.POST.get('postal_code')
            
            # Save the changes
            teacher.user.save()
            teacher.save()
            
            # Handle password change
            current_password = request.POST.get('current_password')
            new_password = request.POST.get('new_password')
            confirm_password = request.POST.get('confirm_password')
            if current_password or new_password or confirm_password:
                if not (current_password and new_password and confirm_password):
                    messages.error(request, 'To change your password, fill in all password fields.')
                    return redirect('teacher_profile_edit')
                if not teacher.user.check_password(current_password):
                    messages.error(request, 'Current password is incorrect.')
                    return redirect('teacher_profile_edit')
                if new_password != confirm_password:
                    messages.error(request, 'New password and confirm password do not match.')
                    return redirect('teacher_profile_edit')
                teacher.user.set_password(new_password)
                teacher.user.save()
                from django.contrib.auth import update_session_auth_hash
                update_session_auth_hash(request, teacher.user)  # Keep user logged in
                messages.success(request, 'Password changed successfully.')
            
            messages.success(request, 'Profile updated successfully.')
            return redirect('teacher_profile')
            
        except Exception as e:
            print("Error updating profile:", str(e))  # Debug print
            messages.error(request, f'Error updating profile: {str(e)}')
    
    context = {
        'teacher': teacher
    }
    return render(request, 'accounts/teacher/profile/edit.html', context)

@login_required
def update_profile_picture(request):
    if request.method == 'POST' and request.FILES.get('photo'):
        profile, created = Profile.objects.get_or_create(user=request.user)
        profile.avatar = request.FILES['photo']
        profile.save()
        messages.success(request, 'Profile picture updated successfully.')
    else:
        messages.error(request, 'Please select a photo to upload.')
    
    return redirect('profile')

@login_required
def student_course_detail(request, course_id):
    if not hasattr(request.user, 'student_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
        
    student = request.user.student_profile
    course = get_object_or_404(Course, id=course_id)
    
    # Check if student is enrolled in this course
    if not student.enrolled_courses.filter(id=course_id).exists():
        messages.error(request, 'You are not enrolled in this course.')
        return redirect('student_courses')
    
    # Get course materials, announcements, assignments and materials
    course_announcements = Announcement.objects.filter(course=course, is_active=True).order_by('-created_at')[:5]
    course_assignments = Activity.objects.filter(course=course, activity_type='assignment').order_by('-due_date')[:5]
    course_exams = Examination.objects.filter(course=course).order_by('-date')[:5]
    course_materials = CourseMaterial.objects.filter(course=course, is_visible=True).order_by('-created_at')
    
    # Debug output
    print(f"Course: {course.title}")
    print(f"Number of materials: {course_materials.count()}")
    for material in course_materials:
        print(f"Material: {material.title} ({material.material_type})")
    
    context = {
        'course': course,
        'student': student,
        'course_announcements': course_announcements,
        'course_assignments': course_assignments,
        'course_exams': course_exams,
        'course_materials': course_materials,
        'today': datetime.now()
    }
    return render(request, 'accounts/student/course_detail.html', context)

@login_required
def student_submit_assignment(request, assignment_id):
    if not hasattr(request.user, 'student_profile'):
        messages.error(request, 'You do not have access to submit assignments.')
        return redirect('student_assignments')
    student = request.user.student_profile
    assignment = get_object_or_404(Activity, id=assignment_id, activity_type='assignment')
    if request.method == 'POST':
        assignment_file = request.FILES.get('assignment_file')
        comments = request.POST.get('comments', '')
        if not assignment_file:
            messages.error(request, 'Please upload a file to submit.')
            return redirect('student_assignments')
        # Save the submission (assume AssignmentSubmission model exists)
        from .models import AssignmentSubmission
        submission, created = AssignmentSubmission.objects.get_or_create(
            student=student, assignment=assignment,
            defaults={'file': assignment_file, 'comments': comments}
        )
        if not created:
            # If already exists, update file and comments
            submission.file = assignment_file
            submission.comments = comments
            submission.save()
        messages.success(request, 'Assignment submitted successfully!')
        return redirect('student_assignments')
    messages.error(request, 'Invalid request method.')
    return redirect('student_assignments')

@login_required
def teacher_course_materials(request):
    """View for listing all course materials for teacher's courses"""
    teacher = request.user.teacher_profile
    courses = Course.objects.filter(teacher=teacher)
    materials = CourseMaterial.objects.filter(course__in=courses).order_by('-created_at')
    
    context = {
        'materials': materials,
        'courses': courses,
        'material_types': CourseMaterial.MATERIAL_TYPES
    }
    return render(request, 'accounts/teacher/course_materials/list.html', context)

@login_required
def teacher_upload_material(request):
    """View for uploading new course material"""
    if request.method == 'POST':
        course_id = request.POST.get('course')
        title = request.POST.get('title')
        description = request.POST.get('description')
        material_type = request.POST.get('material_type')
        file = request.FILES.get('file')
        
        try:
            course = Course.objects.get(id=course_id, teacher=request.user.teacher_profile)
            material = CourseMaterial.objects.create(
                course=course,
                title=title,
                description=description,
                material_type=material_type,
                file=file,
                uploaded_by=request.user
            )
            messages.success(request, 'Course material uploaded successfully!')
            return redirect('teacher_course_materials')
        except Course.DoesNotExist:
            messages.error(request, 'Invalid course selected.')
        except Exception as e:
            messages.error(request, f'Error uploading material: {str(e)}')
    
    return redirect('teacher_course_materials')

@login_required
def teacher_delete_material(request, material_id):
    """View for deleting a course material"""
    try:
        material = CourseMaterial.objects.get(
            id=material_id,
            course__teacher=request.user.teacher_profile
        )
        material.file.delete()  # Delete the actual file
        material.delete()
        messages.success(request, 'Course material deleted successfully!')
    except CourseMaterial.DoesNotExist:
        messages.error(request, 'Material not found or you do not have permission to delete it.')
    except Exception as e:
        messages.error(request, f'Error deleting material: {str(e)}')
    
    return redirect('teacher_course_materials')

@login_required
def student_course_materials(request):
    """View for displaying all course materials for enrolled courses"""
    if not hasattr(request.user, 'student_profile'):
        messages.error(request, 'You do not have access to this page.')
        return redirect('login')
        
    student = request.user.student_profile
    enrolled_courses = student.enrolled_courses.all()
    
    # Get all materials for enrolled courses
    materials_by_course = {}
    for course in enrolled_courses:
        materials = CourseMaterial.objects.filter(
            course=course,
            is_visible=True
        ).order_by('-created_at')
        if materials.exists():
            materials_by_course[course] = materials
    
    context = {
        'materials_by_course': materials_by_course,
        'enrolled_courses': enrolled_courses
    }
    return render(request, 'accounts/student/course_materials.html', context)

@login_required
def export_timetable_pdf(request):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch, cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle
    from datetime import datetime, timedelta, time

    # Get selected class section from query parameters
    selected_class = request.GET.get('class')
    
    # Get all schedules organized by day and time
    schedules = Schedule.objects.select_related('course', 'course__teacher', 'course__class_section').order_by('day', 'start_time')
    
    # Filter schedules by class section if selected
    if selected_class:
        schedules = schedules.filter(course__class_section_id=selected_class)
        selected_class_obj = Class.objects.get(id=selected_class)
        selected_class_name = f"{selected_class_obj.name} - {selected_class_obj.section}"
    else:
        selected_class_name = "All Classes"

    # Create the response object
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="timetable_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'

    # Create the PDF document using landscape orientation with minimal margins
    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        rightMargin=0.5*cm,
        leftMargin=0.5*cm,
        topMargin=0.5*cm,
        bottomMargin=0.5*cm
    )

    # Define styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontSize=12,
        spaceAfter=5
    )
    
    cell_style = ParagraphStyle(
        'CustomCell',
        parent=styles['Normal'],
        fontSize=7,
        leading=8,
        alignment=1
    )

    # Create data for the timetable
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    start_hour = int(request.GET.get('start_hour', 8))
    end_hour = int(request.GET.get('end_hour', 17))
    duration = int(request.GET.get('duration', 60))

    # Create time slots with start and end times
    time_slots = []
    current_hour = start_hour
    while current_hour < end_hour:
        start_time = f"{current_hour:02d}:00"
        next_hour = current_hour + duration // 60
        end_time = f"{next_hour:02d}:00"
        time_slots.append(f"{start_time}-{end_time}")
        current_hour = next_hour

    # Create schedule dictionary for easy lookup
    schedule_dict = {}
    for schedule in schedules:
        start_hour = schedule.start_time.hour
        end_hour = start_hour + duration // 60
        time_key = f"{start_hour:02d}:00-{end_hour:02d}:00"
        day_key = schedule.get_day_display()
        if time_key not in schedule_dict:
            schedule_dict[time_key] = {}
        schedule_dict[time_key][day_key] = schedule

    # Create table data
    table_data = [['Time'] + days]
    for time_slot in time_slots:
        row = [time_slot]
        for day in days:
            cell_content = ''
            if time_slot in schedule_dict and day in schedule_dict[time_slot]:
                schedule = schedule_dict[time_slot][day]
                cell_content = Paragraph(
                    f"<b>{schedule.course.title}</b><br/>"
                    f"{schedule.course.teacher.user.get_full_name()}<br/>"
                    f"<i>R:{schedule.room}</i>",
                    cell_style
                )
            row.append(cell_content)
        table_data.append(row)

    # Calculate optimal dimensions
    page_width, page_height = landscape(A4)
    usable_width = page_width - 1*cm  # Account for margins
    usable_height = page_height - 1*cm  # Account for margins
    
    time_col_width = 2*cm  # Slightly wider for the new time format
    other_col_width = (usable_width - time_col_width) / len(days)
    
    # Create table with optimized dimensions
    table = Table(
        table_data,
        colWidths=[time_col_width] + [other_col_width] * len(days),
        rowHeights=[1*cm] + [1.8*cm] * len(time_slots)
    )

    # Add table style
    table.setStyle(TableStyle([
        # Grid
        ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.black),
        
        # Header row
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        
        # Time column
        ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#f5f5f5')),
        ('FONTSIZE', (0, 1), (0, -1), 7),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica'),
        
        # Alignment
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        
        # Padding
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
    ]))

    # Build the PDF with minimal spacing between elements
    elements = [
        Paragraph(f"Class Timetable - {selected_class_name}", title_style),
        table,
        Paragraph(
            f"Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}", 
            ParagraphStyle('Footer', parent=styles['Normal'], fontSize=6, spaceAfter=0)
        )
    ]
    
    doc.build(elements)
    return response

@login_required
def export_timetable_excel(request):
    # Get selected class section from query parameters
    selected_class = request.GET.get('class')
    
    # Get all schedules organized by day and time
    schedules = Schedule.objects.select_related('course', 'course__teacher', 'course__class_section')
    
    # Filter schedules by class section if selected
    if selected_class:
        schedules = schedules.filter(course__class_section_id=selected_class)
        selected_class_obj = Class.objects.get(id=selected_class)
        selected_class_name = f"{selected_class_obj.name} - {selected_class_obj.section}"
    else:
        selected_class_name = "All Classes"
    
    # Create workbook and active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Timetable"
    
    # Add header
    ws.merge_cells('A1:G1')
    header = ws.cell(row=1, column=1, value=f"Class Timetable - {selected_class_name}")
    header.font = Font(bold=True, size=14)
    header.alignment = Alignment(horizontal='center')
    
    # Add day headers
    days = ['Time', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    header_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
    
    for col, day in enumerate(days, 1):
        cell = ws.cell(row=2, column=col, value=day)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Add time slots and schedule data
    start_hour = int(request.GET.get('start_hour', 8))
    end_hour = int(request.GET.get('end_hour', 17))
    duration = int(request.GET.get('duration', 60))
    
    current_row = 3
    for hour in range(start_hour, end_hour):
        for minute in range(0, 60, duration):
            time_str = f"{hour:02d}:{minute:02d}"
            ws.cell(row=current_row, column=1, value=time_str)
            
            # Add schedule entries for each day
            for day_num, day in enumerate(range(1, 7), 2):  # Start from column 2 (Monday)
                schedule = schedules.filter(
                    day=day,
                    start_time__hour=hour,
                    start_time__minute=minute
                ).first()
                
                if schedule:
                    cell = ws.cell(
                        row=current_row,
                        column=day_num,
                        value=f"{schedule.course.title}\n{schedule.course.teacher.get_full_name() if schedule.course.teacher else ''}\nRoom: {schedule.room}"
                    )
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            current_row += 1
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col[0].column_letter].width = max_length + 2
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="timetable_{selected_class_name.replace(" ", "_")}.xlsx"'
    
    wb.save(response)
    return response

@login_required
def user_manual(request):
    # Render a print-friendly manual; users can export as PDF via browser print
    context = { 'now': timezone.now() }
    return render(request, 'accounts/user_manual.html', context)
